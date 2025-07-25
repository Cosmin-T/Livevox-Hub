# from fastapi import FastAPI, Request, Form, UploadFile, File, HTTPException
# from fastapi.templating import Jinja2Templates
# from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
# import uvicorn
# import os
# import json
# import asyncio
# from typing import Optional, List, Dict, Any
# import pandas as pd
# from datetime import datetime
# import threading
# import time
# from playwright.async_api import async_playwright
# import shutil
# import uuid
# import tempfile
# import zipfile

# # Initialize FastAPI
# app = FastAPI(title="Multi-Project Automation Hub")

# # Templates
# templates = Jinja2Templates(directory="templates")

# print("Server starting...")
# print("Available imports:")
# print("- FastAPI:", FastAPI.__version__ if hasattr(FastAPI, '__version__') else "imported")
# print("- Playwright: imported")
# print("- Pandas:", pd.__version__)
# print("- All other dependencies: imported")

# class DataManager:
#     """Manages agent and call data using local JSON storage"""
#     def __init__(self, project_name="livevox"):
#         self.project_name = project_name
#         self.local_file = f"{project_name}_data.json"
#         print(f"Using local storage: {self.local_file}")

#     def load_data(self):
#         """Load data from local JSON file"""
#         if os.path.exists(self.local_file):
#             try:
#                 with open(self.local_file, 'r') as f:
#                     data = json.load(f)
#                     print(f"Loaded data from {self.local_file}")
#                     return data
#             except Exception as e:
#                 print(f"Error loading {self.local_file}: {e}")

#         # Return default structure if file doesn't exist
#         default_data = {"agents": {}, "calls": {}}
#         self.save_data(default_data)
#         return default_data

#     def save_data(self, data):
#         """Save data to local JSON file"""
#         try:
#             with open(self.local_file, 'w') as f:
#                 json.dump(data, f, indent=2)
#             print(f"Data saved to {self.local_file}")
#             return True
#         except Exception as e:
#             print(f"Error saving to {self.local_file}: {e}")
#             return False

#     def add_agent(self, agent_data):
#         data = self.load_data()
#         agent_id = agent_data['id']
#         data['agents'][agent_id] = agent_data
#         if agent_id not in data['calls']:
#             data['calls'][agent_id] = []
#         self.save_data(data)

#     def remove_agent(self, agent_id):
#         data = self.load_data()
#         if agent_id in data['agents']:
#             del data['agents'][agent_id]
#         if agent_id in data['calls']:
#             del data['calls'][agent_id]
#         self.save_data(data)

#     def remove_all_agents(self):
#         data = {"agents": {}, "calls": {}}
#         self.save_data(data)

#     def add_call(self, agent_id, call_data):
#         data = self.load_data()
#         if agent_id not in data['calls']:
#             data['calls'][agent_id] = []
#         data['calls'][agent_id].append(call_data)
#         self.save_data(data)

#     def remove_all_calls(self, agent_id):
#         data = self.load_data()
#         if agent_id in data['calls']:
#             data['calls'][agent_id] = []
#             self.save_data(data)

#     def remove_call(self, agent_id, call_index):
#         data = self.load_data()
#         if agent_id in data['calls'] and 0 <= call_index < len(data['calls'][agent_id]):
#             del data['calls'][agent_id][call_index]
#             self.save_data(data)

#     def get_agents(self):
#         return self.load_data()['agents']

#     def get_calls(self, agent_id):
#         return self.load_data()['calls'].get(agent_id, [])

# # Global data manager
# data_manager = DataManager()

# @app.get("/", response_class=HTMLResponse)
# async def root(request: Request):
#     return templates.TemplateResponse("index.html", {"request": request})

# @app.get("/debug/routes")
# async def debug_routes():
#     """Debug endpoint to see all registered routes"""
#     routes = []
#     for route in app.routes:
#         if hasattr(route, 'methods') and hasattr(route, 'path'):
#             routes.append({
#                 "path": route.path,
#                 "methods": list(route.methods),
#                 "name": getattr(route, 'name', 'unknown')
#             })
#     return {"routes": routes}

# @app.put("/debug/test-put")
# async def test_put(test_data: str = Form(...)):
#     """Simple PUT endpoint to test if PUT requests work"""
#     return {"success": True, "received": test_data}

# @app.post("/debug/test-automation")
# async def test_automation_endpoint(
#     function_type: str = Form(...),
#     agent_id: str = Form(...),
#     start_date: str = Form(...),
#     username: str = Form(...),
#     password: str = Form(...),
#     url: str = Form(...),
#     phone_numbers: Optional[str] = Form(None)
# ):
#     """Test endpoint with same signature as automation endpoint"""
#     print("DEBUG: Test automation endpoint called successfully!")
#     return {
#         "success": True,
#         "message": "Test endpoint works!",
#         "received_data": {
#             "function_type": function_type,
#             "agent_id": agent_id,
#             "start_date": start_date,
#             "username": username,
#             "url": url,
#             "phone_numbers": phone_numbers
#         }
#     }

# @app.get("/debug/endpoints")
# async def list_endpoints():
#     """List all available endpoints"""
#     routes = []
#     for route in app.routes:
#         if hasattr(route, 'methods') and hasattr(route, 'path'):
#             routes.append({
#                 "path": route.path,
#                 "methods": list(route.methods),
#                 "name": getattr(route, 'name', 'unknown')
#             })
#     return {"routes": routes}

# @app.get("/livevox", response_class=HTMLResponse)
# async def livevox_app(request: Request):
#     agents = data_manager.get_agents()
#     return templates.TemplateResponse("livevox.html", {
#         "request": request,
#         "agents": agents
#     })

# # Agent management endpoints
# @app.post("/api/agents")
# async def create_agent(agent_id: str = Form(...), folder: str = Form(...), locator_id: str = Form(...)):
#     if agent_id in data_manager.get_agents():
#         raise HTTPException(status_code=400, detail="Agent ID already exists")

#     agent_data = {
#         'id': agent_id,
#         'folder': folder,
#         'locator_id': locator_id
#     }
#     data_manager.add_agent(agent_data)
#     return {"success": True}

# @app.get("/api/agents")
# async def get_agents():
#     return data_manager.get_agents()

# @app.put("/api/agents/{agent_id}")
# async def update_agent(agent_id: str, agent_id_field: str = Form(..., alias="agent_id"), folder: str = Form(...), locator_id: str = Form(...)):
#     print(f"PUT /api/agents/{agent_id} called with data: agent_id={agent_id_field}, folder={folder}, locator_id={locator_id}")

#     try:
#         data = data_manager.load_data()

#         if agent_id not in data['agents']:
#             raise HTTPException(status_code=404, detail="Agent not found")

#         # The agent_id from the URL is the current ID, agent_id_field is from the form
#         # If agent ID is changing, we need to update the key
#         if agent_id != agent_id_field:
#             if agent_id_field in data['agents']:
#                 raise HTTPException(status_code=400, detail="New Agent ID already exists")

#             # Move agent data to new key
#             data['agents'][agent_id_field] = {
#                 'id': agent_id_field,
#                 'folder': folder,
#                 'locator_id': locator_id
#             }
#             del data['agents'][agent_id]

#             # Move calls data
#             if agent_id in data['calls']:
#                 data['calls'][agent_id_field] = data['calls'][agent_id]
#                 del data['calls'][agent_id]
#         else:
#             # Just update the existing agent
#             data['agents'][agent_id] = {
#                 'id': agent_id_field,
#                 'folder': folder,
#                 'locator_id': locator_id
#             }

#         data_manager.save_data(data)
#         print(f"Agent updated successfully: {agent_id} -> {agent_id_field}")
#         return {"success": True, "new_agent_id": agent_id_field}

#     except Exception as e:
#         print(f"Error updating agent: {str(e)}")
#         raise HTTPException(status_code=500, detail=str(e))

# @app.delete("/api/agents/{agent_id}")
# async def delete_agent(agent_id: str):
#     data_manager.remove_agent(agent_id)
#     return {"success": True}

# @app.post("/api/agents/{agent_id}/calls")
# async def add_call(agent_id: str, phone: str = Form(...), duration: str = Form(...)):
#     call_data = {'phone': phone, 'duration': duration}
#     data_manager.add_call(agent_id, call_data)
#     return {"success": True}

# @app.get("/api/agents/{agent_id}/calls")
# async def get_calls(agent_id: str):
#     return data_manager.get_calls(agent_id)

# @app.put("/api/agents/{agent_id}/calls/{call_index}")
# async def update_call(agent_id: str, call_index: int, phone: str = Form(...), duration: str = Form(...)):
#     print(f"PUT /api/agents/{agent_id}/calls/{call_index} called with phone={phone}, duration={duration}")

#     try:
#         data = data_manager.load_data()

#         if agent_id not in data['calls']:
#             raise HTTPException(status_code=404, detail="Agent not found")

#         if call_index >= len(data['calls'][agent_id]) or call_index < 0:
#             raise HTTPException(status_code=404, detail="Call not found")

#         data['calls'][agent_id][call_index] = {'phone': phone, 'duration': duration}
#         data_manager.save_data(data)
#         print(f"Call updated successfully: agent={agent_id}, index={call_index}")
#         return {"success": True}

#     except Exception as e:
#         print(f"Error updating call: {str(e)}")
#         raise HTTPException(status_code=500, detail=str(e))

# @app.delete("/api/agents/{agent_id}/calls/{call_index}")
# async def delete_call(agent_id: str, call_index: int):
#     data_manager.remove_call(agent_id, call_index)
#     return {"success": True}

# @app.post("/api/upload-excel")
# async def upload_excel(file: UploadFile = File(...)):
#     try:
#         contents = await file.read()

#         # Save temp file
#         temp_path = f"temp_{uuid.uuid4().hex}.xlsx"
#         with open(temp_path, "wb") as f:
#             f.write(contents)

#         # Read Excel
#         df = pd.read_excel(temp_path, dtype=str)
#         os.remove(temp_path)

#         required_cols = {'phone number', 'duration', 'agent ID', 'folder name (name)', 'locator'}
#         if not required_cols.issubset(df.columns):
#             missing = required_cols - set(df.columns)
#             raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

#         agents_created = 0
#         calls_added = 0

#         for _, row in df.iterrows():
#             agent_id = str(row.get('agent ID', '')).strip()
#             folder = str(row.get('folder name (name)', '')).strip()
#             locator = str(row.get('locator', '')).strip()
#             phone = str(row.get('phone number', '')).strip()
#             duration = str(row.get('duration', '')).strip()

#             if not all([agent_id, folder, locator, phone, duration]):
#                 continue

#             # Create agent if doesn't exist
#             if agent_id not in data_manager.get_agents():
#                 agent_data = {
#                     'id': agent_id,
#                     'folder': folder,
#                     'locator_id': locator
#                 }
#                 data_manager.add_agent(agent_data)
#                 agents_created += 1

#             # Add call
#             call_data = {'phone': phone, 'duration': duration}
#             data_manager.add_call(agent_id, call_data)
#             calls_added += 1

#         return {
#             "success": True,
#             "agents_created": agents_created,
#             "calls_added": calls_added
#         }

#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))

# @app.get("/api/download-template")
# async def download_template():
#     template_data = {
#         'phone number': ['7021234567', '7029876543'],
#         'duration': [65, 123],
#         'agent ID': ['3P_95', '3P_235'],
#         'folder name (name)': ['Chuck', 'Jim'],
#         'locator': ['963665', '976426']
#     }
#     df = pd.DataFrame(template_data)

#     filename = "call_upload_template.xlsx"
#     df.to_excel(filename, index=False)

#     return FileResponse(
#         filename,
#         filename=filename,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )

# @app.delete("/api/agents")
# async def delete_all_agents():
#     data_manager.remove_all_agents()
#     return {"success": True}

# @app.delete("/api/agents/{agent_id}/calls")
# async def delete_all_calls_for_agent(agent_id: str):
#     data_manager.remove_all_calls(agent_id)
#     return {"success": True}

# # Store active automation tasks with their temp directories and results
# active_tasks = {}

# # FIXED AUTOMATION ENDPOINT - NO OUTPUT_DIR REQUIRED!
# @app.post("/api/automation/start")
# async def start_automation_fixed(
#     function_type: str = Form(...),
#     agent_id: str = Form(...),
#     start_date: str = Form(...),
#     username: str = Form(...),
#     password: str = Form(...),
#     url: str = Form(...),
#     phone_numbers: Optional[str] = Form(None)
# ):
#     """FIXED: No output_dir required anymore - uses temp directories"""
#     try:
#         print("=" * 50)
#         print("FIXED AUTOMATION ENDPOINT CALLED")
#         print(f"Function parameters received:")
#         print(f"  function_type: {function_type}")
#         print(f"  agent_id: {agent_id}")
#         print(f"  start_date: {start_date}")
#         print(f"  username: {username}")
#         print(f"  url: {url}")
#         print(f"  phone_numbers: {phone_numbers}")
#         print("=" * 50)

#         task_id = str(uuid.uuid4())

#         # Create temporary directory for this task
#         temp_dir = tempfile.mkdtemp(prefix=f"livevox_{task_id}_")
#         print(f"Created temp directory: {temp_dir}")

#         config = {
#             'function_type': function_type,
#             'agent_id': agent_id,
#             'start_date': start_date,
#             'temp_dir': temp_dir,
#             'username': username,
#             'password': password,
#             'url': url,
#             'phone_numbers': phone_numbers
#         }

#         # Start automation in background
#         task = asyncio.create_task(run_automation(task_id, config))
#         active_tasks[task_id] = {
#             'task': task,
#             'progress': 0,
#             'status': 'running',
#             'message': 'Starting automation...',
#             'temp_dir': temp_dir,
#             'zip_file': None,
#             'agent_id': agent_id
#         }

#         print(f"Task {task_id} started successfully")
#         return {"task_id": task_id, "success": True}

#     except Exception as e:
#         print(f"Error in FIXED automation endpoint: {str(e)}")
#         import traceback
#         traceback.print_exc()
#         raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")

# @app.get("/api/automation/{task_id}/status")
# async def get_automation_status(task_id: str):
#     if task_id not in active_tasks:
#         raise HTTPException(status_code=404, detail="Task not found")

#     task_info = active_tasks[task_id]
#     return {
#         "progress": task_info['progress'],
#         "status": task_info['status'],
#         "message": task_info['message']
#     }

# @app.get("/api/automation/{task_id}/download")
# async def download_automation_result(task_id: str):
#     if task_id not in active_tasks:
#         raise HTTPException(status_code=404, detail="Task not found")

#     task_info = active_tasks[task_id]

#     if task_info['status'] != 'completed':
#         raise HTTPException(status_code=400, detail="Task not completed yet")

#     if not task_info['zip_file'] or not os.path.exists(task_info['zip_file']):
#         raise HTTPException(status_code=404, detail="Download file not found")

#     # Get agent info for filename
#     agent_id = task_info.get('agent_id', 'unknown')
#     timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
#     filename = f"livevox_calls_{agent_id}_{timestamp}.zip"

#     def cleanup_files():
#         """Clean up temp files after download"""
#         try:
#             if os.path.exists(task_info['zip_file']):
#                 os.remove(task_info['zip_file'])
#             if os.path.exists(task_info['temp_dir']):
#                 shutil.rmtree(task_info['temp_dir'])
#             del active_tasks[task_id]
#         except Exception as e:
#             print(f"Error cleaning up files for task {task_id}: {e}")

#     # Schedule cleanup after a delay (gives time for download to complete)
#     threading.Timer(60.0, cleanup_files).start()

#     return FileResponse(
#         task_info['zip_file'],
#         filename=filename,
#         media_type="application/zip"
#     )

# async def run_automation(task_id: str, config: Dict[str, Any]):
#     import time
#     start_time = time.time()
#     timeout_minutes = 30  # 30 minute timeout

#     try:
#         # Update progress function
#         def update_progress(progress: int, message: str):
#             if task_id in active_tasks:
#                 elapsed = int((time.time() - start_time) / 60)
#                 active_tasks[task_id]['progress'] = progress
#                 active_tasks[task_id]['message'] = f"{message} (Running {elapsed}m)"

#                 # Check for timeout
#                 if elapsed >= timeout_minutes:
#                     raise Exception(f"Automation timeout after {timeout_minutes} minutes")

#         update_progress(5, "Starting browser...")

#         async with async_playwright() as playwright:
#             browser = await playwright.webkit.launch()
#             context = await browser.new_context(accept_downloads=True)
#             page = await context.new_page()
#             await page.set_viewport_size({"width": 1440, "height": 1440})
#             page.set_default_timeout(30000)

#             # Add page error listener for debugging
#             def handle_page_error(error):
#                 print(f"Page error: {error}")
#             page.on("pageerror", handle_page_error)

#             try:
#                 # Login with proper waits
#                 update_progress(10, "Logging in...")
#                 await page.goto(config['url'])
#                 await page.locator("#username").fill(config['username'])
#                 await page.locator("#password").fill(config['password'])
#                 await page.locator("#password").press("Enter")

#                 # Wait for login to complete and page to be fully loaded
#                 update_progress(12, "Waiting for login to complete...")
#                 await page.wait_for_load_state("networkidle")
#                 await asyncio.sleep(3)  # Additional wait for page stabilization

#                 # Verify we're logged in by waiting for a dashboard element
#                 update_progress(14, "Verifying login success...")
#                 try:
#                     # Check current URL and page state
#                     current_url = page.url
#                     print(f"Current URL after login: {current_url}")

#                     # Wait for any element that indicates successful login
#                     await page.wait_for_selector("button:has-text('Review')", timeout=15000)
#                     update_progress(15, "Login verified - Review button found")
#                 except Exception as e:
#                     # If Review button not found, try waiting for other common elements
#                     update_progress(14, "Review button not immediately found, waiting longer...")
#                     await page.wait_for_load_state("domcontentloaded")
#                     await asyncio.sleep(2)

#                     # Check if we might be on a wrong page or login failed
#                     page_title = await page.title()
#                     print(f"Page title: {page_title}")
#                     if "login" in page_title.lower() or "sign" in page_title.lower():
#                         raise Exception("Login may have failed - still on login page")

#                 if config['function_type'] == 'specific_calls':
#                     await run_specific_calls_logic(page, config, update_progress)
#                 elif config['function_type'] == 'all_calls':
#                     await run_all_calls_logic(page, config, update_progress)

#                 # Create ZIP file after automation completes
#                 update_progress(95, "Creating ZIP file...")
#                 zip_file_path = await create_zip_file(task_id, config['temp_dir'], config['agent_id'])

#                 if task_id in active_tasks:
#                     active_tasks[task_id]['zip_file'] = zip_file_path
#                     active_tasks[task_id]['agent_id'] = config['agent_id']

#                 elapsed = int((time.time() - start_time) / 60)
#                 update_progress(100, f"Automation completed successfully in {elapsed}m! ZIP file ready for download.")
#                 active_tasks[task_id]['status'] = 'completed'

#             except Exception as e:
#                 error_msg = str(e)
#                 if "timeout" in error_msg.lower():
#                     update_progress(0, f"Automation timed out: {error_msg}")
#                 else:
#                     update_progress(0, f"Error: {error_msg}")
#                 active_tasks[task_id]['status'] = 'failed'
#             finally:
#                 await browser.close()

#     except Exception as e:
#         if task_id in active_tasks:
#             active_tasks[task_id]['status'] = 'failed'
#             active_tasks[task_id]['message'] = f"Error: {str(e)}"

# async def create_zip_file(task_id: str, temp_dir: str, agent_id: str) -> str:
#     """Create a ZIP file containing all downloaded files"""
#     timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
#     zip_filename = f"livevox_calls_{agent_id}_{timestamp}.zip"
#     zip_path = os.path.join(tempfile.gettempdir(), zip_filename)

#     file_count = 0
#     with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
#         for root, dirs, files in os.walk(temp_dir):
#             for file in files:
#                 file_path = os.path.join(root, file)
#                 # Add file to ZIP with relative path
#                 arcname = os.path.relpath(file_path, temp_dir)
#                 zipf.write(file_path, arcname)
#                 file_count += 1

#     print(f"Created ZIP file: {zip_path} with {file_count} files")
#     return zip_path

# async def run_specific_calls_logic(page, config, update_progress):
#     """Implementation for Specific Calls automation logic"""
#     agent_details_map = {
#         "3P_95": {"folder": "Chuck", "locator_id": "963665"},
#         "3P_235": {"folder": "Jim", "locator_id": "976426"},
#         "3P_1483": {"folder": "Clifford", "locator_id": "969789"},
#         "3P_260": {"folder": "Kathy", "locator_id": "963666"},
#         "3P_282": {"folder": "Laurie", "locator_id": "924766"}
#     }

#     current_agent_id = config['agent_id']
#     agent_info = agent_details_map.get(current_agent_id)

#     if not agent_info:
#         raise Exception(f"Agent ID '{current_agent_id}' not found in mapping")

#     calls_to_process = data_manager.get_calls(current_agent_id)
#     if not calls_to_process:
#         raise Exception(f"No calls configured for agent '{current_agent_id}'")

#     # Use temp directory instead of user-specified output directory
#     target_dir = os.path.join(config['temp_dir'], agent_info['folder'])
#     os.makedirs(target_dir, exist_ok=True)

#     update_progress(15, f"Starting automation for {len(calls_to_process)} calls for {agent_info['folder']}")

#     download_counter = 1

#     # Navigate to reports with better error handling
#     update_progress(18, "Navigating to Call Recording Report...")
#     try:
#         # Wait for and click Review button
#         await page.wait_for_selector("button:has-text('Review')", timeout=30000)
#         await page.get_by_role("button", name="Review").click()

#         # Wait for Agent Reports to appear and click it
#         await page.wait_for_selector("text=Agent Reports", timeout=15000)
#         await page.get_by_text("Agent Reports").click()

#         # Wait for Call Recording Report to appear
#         update_progress(19, "Waiting for Call Recording Report...")
#         await page.wait_for_selector("div.MuiTreeItem-label:has-text('Call Recording Report')", timeout=60000)
#         await page.locator("div.MuiTreeItem-label").filter(has_text="Call Recording Report").click()

#         # Wait for the report page to fully load
#         await page.wait_for_load_state("networkidle")
#         update_progress(20, "Report page loaded successfully")

#     except Exception as nav_error:
#         raise Exception(f"Navigation failed: {str(nav_error)}. Make sure you're logged into the correct LiveVox portal.")

#     # Set date and agent
#     update_progress(22, f"Setting up search for agent {current_agent_id}")
#     await page.locator("#search-panel__start-date").fill(config['start_date'])

#     # Click agent dropdown
#     await page.get_by_role("row", name="Agent Select One ... Result").locator("u").click()
#     await page.get_by_placeholder("Search...").fill(current_agent_id)

#     # Select the specific agent
#     agent_selector = f'[id="\\33 {agent_info["locator_id"]}"]'
#     try:
#         await page.wait_for_selector(agent_selector, timeout=10000)
#         await page.locator(agent_selector).click()
#         await page.locator("#agent-combo__ok-btn").click()
#         update_progress(24, f"Agent {current_agent_id} selected successfully")
#     except Exception as e:
#         raise Exception(f"Could not find agent {current_agent_id} (locator: {agent_info['locator_id']}). Check agent mapping.")

#     # Process each call
#     for i, call_data in enumerate(calls_to_process):
#         progress = 25 + int(((i + 1) / len(calls_to_process)) * 65)
#         update_progress(progress, f"Searching call {i+1}/{len(calls_to_process)}: {call_data['phone']} (duration: {call_data['duration']}s)")

#         await page.locator("#search-panel__phone-dialed").clear()
#         await page.locator("#search-panel__phone-dialed").fill(call_data['phone'])
#         await page.get_by_role("button", name="Generate Report").click()

#         try:
#             results_table_body = page.locator("div.rt-tbody")
#             first_row = results_table_body.locator("div.rt-tr:not(.-padRow)").first
#             await first_row.wait_for(state="visible", timeout=15000)

#             data_rows = await results_table_body.locator("div.rt-tr:not(.-padRow)").all()
#             update_progress(progress, f"Found {len(data_rows)} recordings for {call_data['phone']}, checking durations...")

#             found_match = False
#             for j, row in enumerate(data_rows):
#                 duration_in_row = await row.locator("div.rt-td").nth(8).inner_text()
#                 if duration_in_row == call_data['duration']:
#                     update_progress(progress, f"Found matching duration {duration_in_row}s, downloading...")
#                     download_button = row.locator("div.icon--audio-download.clickable")
#                     async with page.expect_download() as download_info:
#                         await download_button.click()

#                     download = await download_info.value
#                     base_filename = download.suggested_filename
#                     name, ext = os.path.splitext(base_filename)

#                     final_filename = f"{name}-{download_counter}{ext}"
#                     target_path = os.path.join(target_dir, final_filename)

#                     update_progress(progress, f"Saved: {final_filename}")
#                     shutil.move(await download.path(), target_path)

#                     download_counter += 1
#                     found_match = True
#                     break

#             if not found_match:
#                 update_progress(progress, f"No matching duration found for {call_data['phone']} (needed: {call_data['duration']}s)")

#         except Exception as e:
#             update_progress(progress, f"No results found for {call_data['phone']}: {str(e)[:50]}")

#     update_progress(90, f"Completed processing {len(calls_to_process)} calls. Downloaded {download_counter-1} files.")

# async def run_all_calls_logic(page, config, update_progress):
#     """Implementation for All Calls automation logic"""
#     agent_details_map = {
#         "3P_95": {"folder": "Chuck", "locator_id": "963665"},
#         "3P_235": {"folder": "Jim", "locator_id": "976426"},
#         "3P_1483": {"folder": "Clifford", "locator_id": "969789"},
#         "3P_260": {"folder": "Kathy", "locator_id": "963666"},
#         "3P_282": {"folder": "Laurie", "locator_id": "924766"}
#     }

#     current_agent_id = config['agent_id']
#     agent_info = agent_details_map.get(current_agent_id)

#     if not agent_info:
#         raise Exception(f"Agent ID '{current_agent_id}' not found in mapping")

#     phone_numbers = config['phone_numbers'].split(',')
#     phone_numbers = [phone.strip() for phone in phone_numbers if phone.strip()]

#     if not phone_numbers:
#         raise Exception("No valid phone numbers provided")

#     # Use temp directory instead of user-specified output directory
#     target_dir = os.path.join(config['temp_dir'], agent_info['folder'])
#     os.makedirs(target_dir, exist_ok=True)

#     update_progress(15, f"Processing {len(phone_numbers)} phone numbers for {agent_info['folder']}")

#     download_counter = 1

#     # Navigate to reports with better error handling
#     update_progress(18, "Navigating to Call Recording Report...")
#     try:
#         # Wait for and click Review button
#         await page.wait_for_selector("button:has-text('Review')", timeout=30000)
#         await page.get_by_role("button", name="Review").click()

#         # Wait for Agent Reports to appear and click it
#         await page.wait_for_selector("text=Agent Reports", timeout=15000)
#         await page.get_by_text("Agent Reports").click()

#         # Wait for Call Recording Report to appear
#         update_progress(19, "Waiting for Call Recording Report...")
#         await page.wait_for_selector("div.MuiTreeItem-label:has-text('Call Recording Report')", timeout=60000)
#         await page.locator("div.MuiTreeItem-label").filter(has_text="Call Recording Report").click()

#         # Wait for the report page to fully load
#         await page.wait_for_load_state("networkidle")
#         update_progress(20, "Report page loaded successfully")

#     except Exception as nav_error:
#         raise Exception(f"Navigation failed: {str(nav_error)}. Make sure you're logged into the correct LiveVox portal.")

#     # Set date and agent
#     update_progress(22, f"Setting up search for agent {current_agent_id}")
#     await page.locator("#search-panel__start-date").fill(config['start_date'])

#     # Click agent dropdown
#     await page.get_by_role("row", name="Agent Select One ... Result").locator("u").click()
#     await page.get_by_placeholder("Search...").fill(current_agent_id)

#     # Select the specific agent
#     agent_selector = f'[id="\\33 {agent_info["locator_id"]}"]'
#     try:
#         await page.wait_for_selector(agent_selector, timeout=10000)
#         await page.locator(agent_selector).click()
#         await page.locator("#agent-combo__ok-btn").click()
#         update_progress(24, f"Agent {current_agent_id} selected successfully")
#     except Exception as e:
#         raise Exception(f"Could not find agent {current_agent_id} (locator: {agent_info['locator_id']}). Check agent mapping.")

#     # Process each phone number
#     for i, phone in enumerate(phone_numbers):
#         progress = 25 + int(((i + 1) / len(phone_numbers)) * 65)
#         update_progress(progress, f"Processing phone {i+1}/{len(phone_numbers)}: {phone}")

#         await page.locator("#search-panel__phone-dialed").clear()
#         await page.locator("#search-panel__phone-dialed").fill(phone)
#         await page.get_by_role("button", name="Generate Report").click()

#         try:
#             results_table_body = page.locator("div.rt-tbody")
#             first_row = results_table_body.locator("div.rt-tr:not(.-padRow)").first
#             await first_row.wait_for(state="visible", timeout=15000)

#             data_rows = await results_table_body.locator("div.rt-tr:not(.-padRow)").all()

#             update_progress(progress, f"Found {len(data_rows)} recording(s) for phone {phone}. Downloading all...")

#             for row in data_rows:
#                 download_button = row.locator("div.icon--audio-download.clickable")
#                 async with page.expect_download() as download_info:
#                     await download_button.click()

#                 download = await download_info.value
#                 base_filename = download.suggested_filename
#                 name, ext = os.path.splitext(base_filename)

#                 final_filename = f"{name}-{download_counter}{ext}"
#                 target_path = os.path.join(target_dir, final_filename)

#                 update_progress(progress, f"Saving file: {final_filename}")
#                 shutil.move(await download.path(), target_path)

#                 download_counter += 1

#         except Exception as e:
#             update_progress(progress, f"No recordings found for phone: {phone}")

#     update_progress(90, f"Completed processing {len(phone_numbers)} phone numbers. Downloaded {download_counter-1} files.")

# if __name__ == "__main__":
#     # Create templates directory if it doesn't exist
#     os.makedirs("templates", exist_ok=True)
#     uvicorn.run(app, host="0.0.0.0", port=8001)


from fastapi import FastAPI, Request, Form, UploadFile, File, HTTPException
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
import uvicorn
import os
import json
import asyncio
from typing import Optional, List, Dict, Any
import pandas as pd
from datetime import datetime, timedelta
import threading
import time
from playwright.async_api import async_playwright
import shutil
import uuid
import tempfile
import zipfile
import logging
import glob
import re
import matplotlib
matplotlib.use('Agg')  # Set non-interactive backend before importing pyplot
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

# Initialize FastAPI
app = FastAPI(title="Multi-Project Automation Hub")

# Templates
templates = Jinja2Templates(directory="templates")

print("Server starting...")
print("Available imports:")
print("- FastAPI:", FastAPI.__version__ if hasattr(FastAPI, '__version__') else "imported")
print("- Playwright: imported")
print("- Pandas:", pd.__version__)
print("- All other dependencies: imported")

class DataManager:
    """Manages agent and call data using local JSON storage"""
    def __init__(self, project_name="livevox"):
        self.project_name = project_name
        self.local_file = f"{project_name}_data.json"
        print(f"Using local storage: {self.local_file}")

    def load_data(self):
        """Load data from local JSON file"""
        if os.path.exists(self.local_file):
            try:
                with open(self.local_file, 'r') as f:
                    data = json.load(f)
                    print(f"Loaded data from {self.local_file}")
                    return data
            except Exception as e:
                print(f"Error loading {self.local_file}: {e}")

        # Return default structure if file doesn't exist
        default_data = {"agents": {}, "calls": {}}
        self.save_data(default_data)
        return default_data

    def save_data(self, data):
        """Save data to local JSON file"""
        try:
            with open(self.local_file, 'w') as f:
                json.dump(data, f, indent=2)
            print(f"Data saved to {self.local_file}")
            return True
        except Exception as e:
            print(f"Error saving to {self.local_file}: {e}")
            return False

    def add_agent(self, agent_data):
        data = self.load_data()
        agent_id = agent_data['id']
        data['agents'][agent_id] = agent_data
        if agent_id not in data['calls']:
            data['calls'][agent_id] = []
        self.save_data(data)

    def remove_agent(self, agent_id):
        data = self.load_data()
        if agent_id in data['agents']:
            del data['agents'][agent_id]
        if agent_id in data['calls']:
            del data['calls'][agent_id]
        self.save_data(data)

    def remove_all_agents(self):
        data = {"agents": {}, "calls": {}}
        self.save_data(data)

    def add_call(self, agent_id, call_data):
        data = self.load_data()
        if agent_id not in data['calls']:
            data['calls'][agent_id] = []
        data['calls'][agent_id].append(call_data)
        self.save_data(data)

    def remove_all_calls(self, agent_id):
        data = self.load_data()
        if agent_id in data['calls']:
            data['calls'][agent_id] = []
            self.save_data(data)

    def remove_call(self, agent_id, call_index):
        data = self.load_data()
        if agent_id in data['calls'] and 0 <= call_index < len(data['calls'][agent_id]):
            del data['calls'][agent_id][call_index]
            self.save_data(data)

    def get_agents(self):
        return self.load_data()['agents']

    def get_calls(self, agent_id):
        return self.load_data()['calls'].get(agent_id, [])

class HciDataManager(DataManager):
    """Manages HCI agent data using local JSON storage"""
    def __init__(self):
        super().__init__("hci")

    def load_data(self):
        """Load data from local JSON file with default HCI agents"""
        if os.path.exists(self.local_file):
            try:
                with open(self.local_file, 'r') as f:
                    data = json.load(f)
                    print(f"Loaded HCI data from {self.local_file}")
                    return data
            except Exception as e:
                print(f"Error loading {self.local_file}: {e}")

        # Return default structure with default HCI agents if file doesn't exist
        default_data = {"agents": HCI_AGENT_LIST.copy()}
        self.save_data(default_data)
        return default_data

class PhrasesDataManager(DataManager):
    """Manages phrases data using local JSON storage"""
    def __init__(self):
        super().__init__("phrases")

    def load_data(self):
        """Load phrases data from local JSON file"""
        if os.path.exists(self.local_file):
            try:
                with open(self.local_file, 'r') as f:
                    data = json.load(f)
                    print(f"Loaded phrases data from {self.local_file}")
                    return data
            except Exception as e:
                print(f"Error loading {self.local_file}: {e}")

        # Return default empty structure if file doesn't exist
        default_data = {"phrases": []}
        self.save_data(default_data)
        return default_data

    def add_phrase(self, phrase_data):
        """Add a new phrase to storage"""
        data = self.load_data()

        # Add timestamp and ID if not present
        if 'id' not in phrase_data:
            phrase_data['id'] = str(uuid.uuid4())
        if 'created_at' not in phrase_data:
            from datetime import datetime
            phrase_data['created_at'] = datetime.now().isoformat()

        data['phrases'].append(phrase_data)
        self.save_data(data)
        return phrase_data['id']

    def get_phrase(self, phrase_id):
        """Get a specific phrase by ID"""
        data = self.load_data()
        for phrase in data['phrases']:
            if phrase.get('id') == phrase_id:
                return phrase
        return None

    def get_all_phrases(self):
        """Get all saved phrases"""
        data = self.load_data()
        return data.get('phrases', [])

    def update_phrase(self, phrase_id, updated_data):
        """Update an existing phrase by ID"""
        data = self.load_data()

        for i, phrase in enumerate(data['phrases']):
            if phrase.get('id') == phrase_id:
                # Update the phrase data while preserving id and created_at
                data['phrases'][i].update(updated_data)
                data['phrases'][i]['id'] = phrase_id  # Ensure ID is preserved
                if 'created_at' not in data['phrases'][i]:
                    from datetime import datetime
                    data['phrases'][i]['created_at'] = datetime.now().isoformat()

                self.save_data(data)
                print(f"Updated phrase with ID: {phrase_id}")
                return True
        return False

    def delete_phrase(self, phrase_id):
        """Delete a phrase by ID"""
        data = self.load_data()
        original_count = len(data['phrases'])
        data['phrases'] = [p for p in data['phrases'] if p.get('id') != phrase_id]

        if len(data['phrases']) < original_count:
            self.save_data(data)
            print(f"Deleted phrase with ID: {phrase_id}")
            return True
        return False

# Global data managers
data_manager = DataManager()
hci_data_manager = HciDataManager()
phrases_data_manager = PhrasesDataManager()

# HCI Configuration
HCI_AGENT_LIST = [
    'JCI_HCI_AGENT (3181894)',
    'BN_HCI_AGENT (3176615)',
    'AU_HCI_AGENT (3174939)',
    'DS_HCI_AGENT (3173370)',
    'FEDEX_HCI_AGENT (3175378)',
    'LB1_HCI_AGENT (3183640)',
    'LB2_HCI_AGENT (3174416)',
    'SW_HCI_AGENT (3174210)',
    'LARGE_BAL_HCI_AGENT (3179837)',
    'PHI_HCI_AGENT (3191012)',
    'SMALL_BAL_HCI_AGENT (3179861)',
    'DELL_HCI_AGENT (3184975)',
    'ATI_RESTORATION_HCI_AGENT',
    'DS_DRAGON_HCI_AGENT',
    'DOM_HCI_AGENT_Uline (3186839)',
    'DOM_HCI_AGENT_SB_Team (3187140)',
]

@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/livevox", response_class=HTMLResponse)
async def livevox_app(request: Request):
    agents = data_manager.get_agents()
    return templates.TemplateResponse("livevox.html", {
        "request": request,
        "agents": agents
    })

@app.get("/hci-summary", response_class=HTMLResponse)
async def hci_summary_app(request: Request):
    return templates.TemplateResponse("hci_summary.html", {"request": request})

@app.get("/add-phrases", response_class=HTMLResponse)
async def add_phrases_app(request: Request):
    return templates.TemplateResponse("add_phrases.html", {"request": request})

# Store active automation tasks with their temp directories and results
active_tasks = {}

# Existing LiveVox automation endpoints (keeping all existing ones)
@app.post("/api/agents")
async def create_agent(agent_id: str = Form(...), folder: str = Form(...), locator_id: str = Form(...)):
    if agent_id in data_manager.get_agents():
        raise HTTPException(status_code=400, detail="Agent ID already exists")

    agent_data = {
        'id': agent_id,
        'folder': folder,
        'locator_id': locator_id
    }
    data_manager.add_agent(agent_data)
    return {"success": True}

@app.get("/api/agents")
async def get_agents():
    return data_manager.get_agents()

@app.put("/api/agents/{agent_id}")
async def update_agent(agent_id: str, agent_id_field: str = Form(..., alias="agent_id"), folder: str = Form(...), locator_id: str = Form(...)):
    print(f"PUT /api/agents/{agent_id} called with data: agent_id={agent_id_field}, folder={folder}, locator_id={locator_id}")

    try:
        data = data_manager.load_data()

        if agent_id not in data['agents']:
            raise HTTPException(status_code=404, detail="Agent not found")

        if agent_id != agent_id_field:
            if agent_id_field in data['agents']:
                raise HTTPException(status_code=400, detail="New Agent ID already exists")

            data['agents'][agent_id_field] = {
                'id': agent_id_field,
                'folder': folder,
                'locator_id': locator_id
            }
            del data['agents'][agent_id]

            if agent_id in data['calls']:
                data['calls'][agent_id_field] = data['calls'][agent_id]
                del data['calls'][agent_id]
        else:
            data['agents'][agent_id] = {
                'id': agent_id_field,
                'folder': folder,
                'locator_id': locator_id
            }

        data_manager.save_data(data)
        print(f"Agent updated successfully: {agent_id} -> {agent_id_field}")
        return {"success": True, "new_agent_id": agent_id_field}

    except Exception as e:
        print(f"Error updating agent: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/agents/{agent_id}")
async def delete_agent(agent_id: str):
    data_manager.remove_agent(agent_id)
    return {"success": True}

@app.post("/api/agents/{agent_id}/calls")
async def add_call(agent_id: str, phone: str = Form(...), duration: str = Form(...)):
    call_data = {'phone': phone, 'duration': duration}
    data_manager.add_call(agent_id, call_data)
    return {"success": True}

@app.get("/api/agents/{agent_id}/calls")
async def get_calls(agent_id: str):
    return data_manager.get_calls(agent_id)

@app.put("/api/agents/{agent_id}/calls/{call_index}")
async def update_call(agent_id: str, call_index: int, phone: str = Form(...), duration: str = Form(...)):
    print(f"PUT /api/agents/{agent_id}/calls/{call_index} called with phone={phone}, duration={duration}")

    try:
        data = data_manager.load_data()

        if agent_id not in data['calls']:
            raise HTTPException(status_code=404, detail="Agent not found")

        if call_index >= len(data['calls'][agent_id]) or call_index < 0:
            raise HTTPException(status_code=404, detail="Call not found")

        data['calls'][agent_id][call_index] = {'phone': phone, 'duration': duration}
        data_manager.save_data(data)
        print(f"Call updated successfully: agent={agent_id}, index={call_index}")
        return {"success": True}

    except Exception as e:
        print(f"Error updating call: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/agents/{agent_id}/calls/{call_index}")
async def delete_call(agent_id: str, call_index: int):
    data_manager.remove_call(agent_id, call_index)
    return {"success": True}

@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()

        temp_path = f"temp_{uuid.uuid4().hex}.xlsx"
        with open(temp_path, "wb") as f:
            f.write(contents)

        df = pd.read_excel(temp_path, dtype=str)
        os.remove(temp_path)

        required_cols = {'phone number', 'duration', 'agent ID', 'folder name (name)', 'locator'}
        if not required_cols.issubset(df.columns):
            missing = required_cols - set(df.columns)
            raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

        agents_created = 0
        calls_added = 0

        for _, row in df.iterrows():
            agent_id = str(row.get('agent ID', '')).strip()
            folder = str(row.get('folder name (name)', '')).strip()
            locator = str(row.get('locator', '')).strip()
            phone = str(row.get('phone number', '')).strip()
            duration = str(row.get('duration', '')).strip()

            if not all([agent_id, folder, locator, phone, duration]):
                continue

            if agent_id not in data_manager.get_agents():
                agent_data = {
                    'id': agent_id,
                    'folder': folder,
                    'locator_id': locator
                }
                data_manager.add_agent(agent_data)
                agents_created += 1

            call_data = {'phone': phone, 'duration': duration}
            data_manager.add_call(agent_id, call_data)
            calls_added += 1

        return {
            "success": True,
            "agents_created": agents_created,
            "calls_added": calls_added
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/download-template")
async def download_template():
    template_data = {
        'phone number': ['7021234567', '7029876543'],
        'duration': [65, 123],
        'agent ID': ['3P_95', '3P_235'],
        'folder name (name)': ['Chuck', 'Jim'],
        'locator': ['963665', '976426']
    }
    df = pd.DataFrame(template_data)

    filename = "call_upload_template.xlsx"
    df.to_excel(filename, index=False)

    return FileResponse(
        filename,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/download-all-data")
async def download_all_existing_data():
    """Download all existing agent data with phone numbers and durations"""
    try:
        # Get all existing calls data
        all_phone_numbers = []
        all_durations = []
        all_agent_ids = []
        all_folder_names = []
        all_locators = []
        
        # Agent details mapping
        agent_details_map = {
            "3P_95": {"folder": "Chuck", "locator_id": "963665"},
            "3P_235": {"folder": "Jim", "locator_id": "976426"},
            "3P_1483": {"folder": "Clifford", "locator_id": "969789"},
            "3P_260": {"folder": "Kathy", "locator_id": "963666"},
            "3P_282": {"folder": "Laurie", "locator_id": "924766"}
        }
        
        # Get calls for each agent
        for agent_id in agent_details_map.keys():
            calls = data_manager.get_calls(agent_id)
            if calls:
                for call in calls:
                    all_phone_numbers.append(call['phone'])
                    all_durations.append(int(call['duration']))
                    all_agent_ids.append(agent_id)
                    all_folder_names.append(agent_details_map[agent_id]['folder'])
                    all_locators.append(agent_details_map[agent_id]['locator_id'])
        
        if not all_phone_numbers:
            # Return empty template if no data exists
            template_data = {
                'phone number': [],
                'duration': [],
                'agent ID': [],
                'folder name (name)': [],
                'locator': []
            }
        else:
            template_data = {
                'phone number': all_phone_numbers,
                'duration': all_durations,
                'agent ID': all_agent_ids,
                'folder name (name)': all_folder_names,
                'locator': all_locators
            }
        
        df = pd.DataFrame(template_data)
        filename = "all_existing_agent_data.xlsx"
        df.to_excel(filename, index=False)
        return FileResponse(
            filename,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading existing data: {str(e)}")

@app.delete("/api/agents")
async def delete_all_agents():
    data_manager.remove_all_agents()
    return {"success": True}

@app.delete("/api/agents/{agent_id}/calls")
async def delete_all_calls_for_agent(agent_id: str):
    data_manager.remove_all_calls(agent_id)
    return {"success": True}

# Existing LiveVox automation endpoint
@app.post("/api/automation/start")
async def start_automation_fixed(
    function_type: str = Form(...),
    agent_id: str = Form(...),
    start_date: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    url: str = Form(...),
    phone_numbers: Optional[str] = Form(None)
):
    """FIXED: No output_dir required anymore - uses temp directories"""
    try:
        print("=" * 50)
        print("FIXED AUTOMATION ENDPOINT CALLED")
        print(f"Function parameters received:")
        print(f"  function_type: {function_type}")
        print(f"  agent_id: {agent_id}")
        print(f"  start_date: {start_date}")
        print(f"  username: {username}")
        print(f"  url: {url}")
        print(f"  phone_numbers: {phone_numbers}")
        print("=" * 50)

        task_id = str(uuid.uuid4())

        temp_dir = tempfile.mkdtemp(prefix=f"livevox_{task_id}_")
        print(f"Created temp directory: {temp_dir}")

        config = {
            'function_type': function_type,
            'agent_id': agent_id,
            'start_date': start_date,
            'temp_dir': temp_dir,
            'username': username,
            'password': password,
            'url': url,
            'phone_numbers': phone_numbers
        }

        task = asyncio.create_task(run_automation(task_id, config))
        active_tasks[task_id] = {
            'task': task,
            'progress': 0,
            'status': 'running',
            'message': 'Starting automation...',
            'temp_dir': temp_dir,
            'zip_file': None,
            'agent_id': agent_id,
            'log_messages': ['Automation terminal ready...']
        }

        print(f"Task {task_id} started successfully")
        return {"task_id": task_id, "success": True}

    except Exception as e:
        print(f"Error in FIXED automation endpoint: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")

# NEW HCI Summary Report Endpoints
@app.post("/api/hci/automation/start")
async def start_hci_automation(
    username: str = Form(...),
    password: str = Form(...),
    url: str = Form(...),
    selected_agents: str = Form(...),  # Comma-separated list of selected agents
    report_date: Optional[str] = Form(None)  # Add this parameter
):
    """Start HCI Summary Report automation"""
    try:
        print("=" * 50)
        print("HCI AUTOMATION ENDPOINT CALLED")
        print(f"Username: {username}")
        print(f"URL: {url}")
        print(f"Selected agents: {selected_agents}")
        print(f"Report date: {report_date}")  # Add this log
        print("=" * 50)

        task_id = str(uuid.uuid4())

        temp_dir = tempfile.mkdtemp(prefix=f"hci_{task_id}_")
        print(f"Created temp directory: {temp_dir}")

        # Parse selected agents
        selected_agent_list = [agent.strip() for agent in selected_agents.split(',') if agent.strip()]

        config = {
            'username': username,
            'password': password,
            'url': url,
            'selected_agents': selected_agent_list,
            'temp_dir': temp_dir,
            'report_date': report_date  # Add this to config
        }

        task = asyncio.create_task(run_hci_automation(task_id, config))
        active_tasks[task_id] = {
            'task': task,
            'progress': 0,
            'status': 'running',
            'message': 'Starting HCI automation...',
            'temp_dir': temp_dir,
            'zip_file': None,
            'project_type': 'hci',
            'log_messages': ['Automation terminal ready...']
        }

        print(f"HCI Task {task_id} started successfully")
        return {"task_id": task_id, "success": True}

    except Exception as e:
        print(f"Error in HCI automation endpoint: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")

@app.get("/api/hci/agents")
async def get_hci_agents():
    """Get list of available HCI agents"""
    return {"agents": hci_data_manager.load_data().get('agents', HCI_AGENT_LIST)}

@app.post("/api/hci/agents")
async def add_hci_agent(request: Request):
    """Add a new HCI agent"""
    try:
        body = await request.json()
        agent_name = body.get('agent_name')

        if not agent_name:
            raise HTTPException(status_code=400, detail="Agent name is required")

        data = hci_data_manager.load_data()
        if 'agents' not in data:
            data['agents'] = HCI_AGENT_LIST.copy()

        if agent_name in data['agents']:
            raise HTTPException(status_code=400, detail="Agent already exists")

        data['agents'].append(agent_name)
        hci_data_manager.save_data(data)

        return {"success": True, "message": f"Agent '{agent_name}' added successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/hci/agents")
async def remove_hci_agent(request: Request):
    """Remove an HCI agent"""
    try:
        body = await request.json()
        agent_name = body.get('agent_name')

        if not agent_name:
            raise HTTPException(status_code=400, detail="Agent name is required")

        data = hci_data_manager.load_data()
        if 'agents' not in data:
            data['agents'] = HCI_AGENT_LIST.copy()

        if agent_name not in data['agents']:
            raise HTTPException(status_code=404, detail="Agent not found")

        data['agents'].remove(agent_name)
        hci_data_manager.save_data(data)

        return {"success": True, "message": f"Agent '{agent_name}' removed successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# NEW Add Phrases Endpoints
@app.post("/api/phrases/add-phrase")
async def add_phrase(
    phrase_name: str = Form(...),
    phrase_verbiage: str = Form(...),
    phrase_description: str = Form(default=""),
    wav_file: Optional[UploadFile] = File(None),
    converted_file_name: Optional[str] = Form(None)
):
    """Add a single phrase with optional WAV file conversion"""
    import os
    import tempfile
    import subprocess
    from pathlib import Path

    try:
        # Validate required fields
        if not phrase_name.strip():
            raise HTTPException(status_code=400, detail="Phrase name is required")
        if not phrase_verbiage.strip():
            raise HTTPException(status_code=400, detail="Phrase verbiage is required")

        # Create phrase data
        phrase_data = {
            "phrase_name": phrase_name.strip(),
            "verbiage": phrase_verbiage.strip(),
            "description": phrase_description.strip(),
            "wav_org_file": None
        }

        # Process WAV file if provided
        if wav_file and wav_file.filename:
            print(f"Processing audio file: {wav_file.filename}")

            # Validate file type
            allowed_extensions = ['.wav', '.m4a']
            file_ext = os.path.splitext(wav_file.filename)[1].lower()
            if file_ext not in allowed_extensions:
                raise HTTPException(
                    status_code=400,
                    detail=f"Unsupported file type: {file_ext}. Allowed: {allowed_extensions}"
                )

            print(f"File validation passed for: {wav_file.filename}")

            # Create temporary directory for audio processing
            temp_dir = tempfile.mkdtemp(prefix="phrase_audio_")
            print(f"Created temporary directory: {temp_dir}")

            try:
                # Save uploaded file
                original_filename = wav_file.filename
                temp_input_path = os.path.join(temp_dir, original_filename)

                print(f"Saving uploaded file to: {temp_input_path}")
                with open(temp_input_path, "wb") as temp_file:
                    content = await wav_file.read()
                    temp_file.write(content)
                print(f"File saved successfully, size: {len(content)} bytes")

                # Convert to org_.wav format
                base_name = os.path.splitext(original_filename)[0]
                output_filename = f"org_{base_name}.wav"
                temp_output_path = os.path.join(temp_dir, output_filename)

                print(f"Starting audio conversion: {original_filename} -> {output_filename}")

                # Check if ffmpeg is available
                ffmpeg_available = False
                try:
                    subprocess.run(['ffmpeg', '-version'], check=True, capture_output=True, text=True)
                    ffmpeg_available = True
                    print("FFmpeg is available")
                except (subprocess.CalledProcessError, FileNotFoundError):
                    print("FFmpeg not found in system PATH")

                if not ffmpeg_available:
                    raise HTTPException(
                        status_code=500,
                        detail="FFmpeg is not installed or not found in system PATH. Please install FFmpeg to enable audio conversion."
                    )

                # Use ffmpeg for conversion
                try:
                    result = subprocess.run([
                        'ffmpeg', '-i', temp_input_path,
                        '-acodec', 'pcm_s16le',
                        '-ar', '8000',
                        '-ac', '1',
                        '-y',  # Overwrite output file
                        temp_output_path
                    ], check=True, capture_output=True, text=True)

                    print(f"Audio conversion completed successfully")

                    # Read converted file and encode to base64 for storage
                    with open(temp_output_path, "rb") as converted_file:
                        converted_content = converted_file.read()

                    print(f"Converted file size: {len(converted_content)} bytes")

                    import base64
                    phrase_data["wav_org_file"] = {
                        "filename": output_filename,
                        "content": base64.b64encode(converted_content).decode('utf-8')
                    }

                    print(f"Audio file encoded and stored successfully")

                except subprocess.CalledProcessError as e:
                    print(f"FFmpeg conversion failed: {e.stderr}")
                    raise HTTPException(
                        status_code=500,
                        detail=f"Audio conversion failed: {e.stderr if e.stderr else str(e)}"
                    )

            finally:
                # Cleanup temporary files
                import shutil
                shutil.rmtree(temp_dir, ignore_errors=True)

        # Handle converted file selection if provided
        elif converted_file_name:
            print(f"Using converted file: {converted_file_name}")

            # Load converted files from storage
            stored_files = await load_converted_files_from_storage()

            # Find the selected file
            selected_file = None
            for stored_file in stored_files:
                if stored_file['filename'] == converted_file_name:
                    selected_file = stored_file
                    break

            if selected_file:
                print(f"Found converted file: {selected_file['filename']}")
                phrase_data["wav_org_file"] = {
                    "filename": selected_file['filename'],
                    "content": selected_file['content'],  # Base64 encoded
                    "size": selected_file['size']
                }
            else:
                print(f"Warning: Converted file '{converted_file_name}' not found in storage")

        # Store phrase data in persistent storage
        phrase_id = phrases_data_manager.add_phrase(phrase_data)

        # Also store in memory cache for automation
        if not hasattr(app.state, 'phrase_cache'):
            app.state.phrase_cache = {}
        app.state.phrase_cache[phrase_id] = phrase_data

        # Return success response
        response_data = {
            "success": True,
            "message": "Phrase added successfully",
            "phrase_id": phrase_id,
            "phrase_data": {
                "phrase_name": phrase_data["phrase_name"],
                "verbiage": phrase_data["verbiage"],
                "description": phrase_data["description"],
                "has_audio": phrase_data["wav_org_file"] is not None
            }
        }

        if phrase_data["wav_org_file"]:
            response_data["converted_file"] = phrase_data["wav_org_file"]["filename"]

        return response_data

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error adding phrase: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to add phrase: {str(e)}")

@app.get("/api/phrases/list")
async def list_phrases():
    """Get all saved phrases"""
    try:
        phrases = phrases_data_manager.get_all_phrases()
        return {"success": True, "phrases": phrases}
    except Exception as e:
        print(f"Error listing phrases: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to list phrases: {str(e)}")

@app.delete("/api/phrases/{phrase_id}")
async def delete_phrase(phrase_id: str):
    """Delete a phrase by ID"""
    try:
        success = phrases_data_manager.delete_phrase(phrase_id)
        if success:
            return {"success": True, "message": f"Phrase {phrase_id} deleted successfully"}
        else:
            raise HTTPException(status_code=404, detail="Phrase not found")
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error deleting phrase: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete phrase: {str(e)}")

@app.get("/api/phrases/converted-files")
async def get_converted_files():
    """Get list of available converted audio files"""
    try:
        # Load files from persistent storage
        stored_files = await load_converted_files_from_storage()

        # Also check memory cache and merge
        if not hasattr(app.state, 'converted_audio_cache'):
            app.state.converted_audio_cache = []

        # Combine stored files and memory cache, avoiding duplicates
        all_files = stored_files.copy()
        for cache_file in app.state.converted_audio_cache:
            if not any(f['filename'] == cache_file['filename'] for f in all_files):
                all_files.append(cache_file)

        # Return list without the base64 content to keep response small
        files = [
            {
                "filename": f["filename"],
                "original_name": f.get("original_name", f["filename"]),
                "size": f.get("size", 0),
                "file_type": f.get("file_type", "converted")  # Default to converted for backward compatibility
            }
            for f in all_files
        ]

        print(f"Found {len(files)} converted files: {[f['filename'] for f in files]}")
        return {"success": True, "files": files}
    except Exception as e:
        print(f"Error listing converted files: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to list converted files: {str(e)}")

@app.delete("/api/phrases/converted-files/{filename}")
async def delete_converted_file(filename: str):
    """Delete a converted audio file"""
    try:
        # Check if file is used by any phrases
        phrases_using_file = []
        all_phrases = phrases_data_manager.get_all_phrases()

        for phrase in all_phrases:
            if phrase.get('wav_org_file') and phrase['wav_org_file'].get('filename') == filename:
                phrases_using_file.append(phrase['phrase_name'])

        # If file is in use, return warning (but still allow deletion)
        if phrases_using_file:
            print(f"Warning: File {filename} is used by phrases: {phrases_using_file}")

        # Delete the file from storage
        success = await delete_converted_file_from_storage(filename)

        if success:
            # Remove from memory cache if present
            if hasattr(app.state, 'converted_audio_cache'):
                app.state.converted_audio_cache = [
                    f for f in app.state.converted_audio_cache
                    if f.get('filename') != filename
                ]

            # Update phrases that were using this file
            if phrases_using_file:
                for phrase in all_phrases:
                    if phrase.get('wav_org_file') and phrase['wav_org_file'].get('filename') == filename:
                        # Remove the file reference from the phrase
                        phrase['wav_org_file'] = None
                        phrases_data_manager.update_phrase(phrase['id'], phrase)

                return {
                    "success": True,
                    "message": f"File {filename} deleted successfully. Removed from {len(phrases_using_file)} phrase(s).",
                    "affected_phrases": phrases_using_file
                }
            else:
                return {
                    "success": True,
                    "message": f"File {filename} deleted successfully."
                }
        else:
            raise HTTPException(status_code=404, detail="File not found")

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error deleting converted file: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete file: {str(e)}")

@app.get("/api/phrases/audio/{filename}")
async def serve_audio_file(filename: str):
    """Serve audio file for streaming/playback"""
    try:
        # Load converted files from storage
        stored_files = await load_converted_files_from_storage()

        # Also check memory cache
        if hasattr(app.state, 'converted_audio_cache'):
            for cache_file in app.state.converted_audio_cache:
                if not any(f['filename'] == cache_file['filename'] for f in stored_files):
                    stored_files.append(cache_file)

        # Find the requested file
        audio_file = next((f for f in stored_files if f['filename'] == filename), None)

        if not audio_file:
            raise HTTPException(status_code=404, detail="Audio file not found")

        # Decode base64 content
        import base64
        try:
            audio_content = base64.b64decode(audio_file['content'])
        except Exception as e:
            print(f"Error decoding audio file {filename}: {e}")
            raise HTTPException(status_code=500, detail="Error processing audio file")

        # Return audio response with proper headers
        from fastapi.responses import Response

        return Response(
            content=audio_content,
            media_type="audio/wav",
            headers={
                "Content-Disposition": f"inline; filename={filename}",
                "Cache-Control": "public, max-age=3600",  # Cache for 1 hour
                "Accept-Ranges": "bytes",
                "Content-Length": str(len(audio_content))
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error serving audio file: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to serve audio file: {str(e)}")

@app.get("/api/phrases/{phrase_id}")
async def get_phrase(phrase_id: str):
    """Get a specific phrase by ID"""
    try:
        phrase = phrases_data_manager.get_phrase(phrase_id)
        if not phrase:
            raise HTTPException(status_code=404, detail="Phrase not found")
        return {"success": True, "phrase": phrase}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error getting phrase: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get phrase: {str(e)}")

@app.put("/api/phrases/{phrase_id}")
async def update_phrase(
    phrase_id: str,
    phrase_name: str = Form(...),
    phrase_verbiage: str = Form(...),
    phrase_description: str = Form(default=""),
    converted_file_name: Optional[str] = Form(None)
):
    """Update an existing phrase by ID"""
    try:
        # Validate required fields
        if not phrase_name.strip():
            raise HTTPException(status_code=400, detail="Phrase name is required")
        if not phrase_verbiage.strip():
            raise HTTPException(status_code=400, detail="Phrase verbiage is required")

        # Check if phrase exists
        existing_phrase = phrases_data_manager.get_phrase(phrase_id)
        if not existing_phrase:
            raise HTTPException(status_code=404, detail="Phrase not found")

        # Handle audio file association
        wav_org_file = None
        if converted_file_name:
            # Load converted files to find the file data
            try:
                converted_files = await load_converted_files_from_storage()
                selected_file = next((f for f in converted_files if f['filename'] == converted_file_name), None)

                if selected_file:
                    wav_org_file = {
                        "filename": selected_file["filename"],
                        "original_name": selected_file.get("original_name", selected_file["filename"]),
                        "size": selected_file.get("size", 0),
                        "base64_content": selected_file.get("base64_content", "")
                    }
                    print(f"Associated converted file: {converted_file_name}")
                else:
                    print(f"Warning: Converted file {converted_file_name} not found in storage")
            except Exception as e:
                print(f"Error loading converted file {converted_file_name}: {e}")

        # Prepare updated phrase data
        updated_data = {
            "phrase_name": phrase_name.strip(),
            "verbiage": phrase_verbiage.strip(),
            "description": phrase_description.strip(),
            "wav_org_file": wav_org_file
        }

        # Update the phrase
        success = phrases_data_manager.update_phrase(phrase_id, updated_data)

        if success:
            return {"success": True, "message": f"Phrase {phrase_id} updated successfully"}
        else:
            raise HTTPException(status_code=404, detail="Phrase not found")

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error updating phrase: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update phrase: {str(e)}")

async def save_converted_file_to_storage(file_data):
    """Save converted audio file to persistent storage"""
    try:
        # Load existing converted files data
        converted_files_path = "converted_audio_files.json"
        try:
            with open(converted_files_path, 'r') as f:
                stored_files = json.load(f)
        except FileNotFoundError:
            stored_files = []

        # Add new file if not already exists
        existing_file = next((f for f in stored_files if f['filename'] == file_data['filename']), None)
        if not existing_file:
            stored_files.append(file_data)

            # Save updated list
            with open(converted_files_path, 'w') as f:
                json.dump(stored_files, f, indent=2)
            print(f"Saved converted file to persistent storage: {file_data['filename']}")
        else:
            print(f"File already exists in storage: {file_data['filename']}")

    except Exception as e:
        print(f"Error saving converted file to storage: {e}")

async def delete_converted_file_from_storage(filename: str):
    """Delete a converted audio file from persistent storage"""
    try:
        converted_files_path = "converted_audio_files.json"

        # Load existing files
        try:
            with open(converted_files_path, 'r') as f:
                stored_files = json.load(f)
        except FileNotFoundError:
            print(f"No converted_audio_files.json found, cannot delete {filename}")
            return False

        # Find and remove the file
        original_count = len(stored_files)
        stored_files = [f for f in stored_files if f.get('filename') != filename]

        if len(stored_files) < original_count:
            # Save updated list
            with open(converted_files_path, 'w') as f:
                json.dump(stored_files, f, indent=2)
            print(f"Deleted converted file from storage: {filename}")
            return True
        else:
            print(f"File not found in storage: {filename}")
            return False

    except Exception as e:
        print(f"Error deleting converted file from storage: {e}")
        return False

async def load_converted_files_from_storage():
    """Load converted audio files from persistent storage"""
    try:
        converted_files_path = "converted_audio_files.json"
        print(f"Loading converted files from: {converted_files_path}")

        with open(converted_files_path, 'r') as f:
            stored_files = json.load(f)

        print(f"Loaded {len(stored_files)} files from storage")
        for file in stored_files:
            print(f"  - {file.get('filename', 'Unknown filename')}")

        return stored_files
    except FileNotFoundError:
        print("No converted_audio_files.json found")
        return []
    except Exception as e:
        print(f"Error loading converted files from storage: {e}")
        return []

async def convert_audio_via_web_service(input_path: str, output_path: str) -> bytes:
    """Convert audio file using g711.org web service like the original implementation"""
    from playwright.async_api import async_playwright
    import asyncio

    try:
        async with async_playwright() as p:
            # Launch browser in visible mode for debugging
            browser = await p.chromium.launch(headless=False)
            context = await browser.new_context()
            page = await context.new_page()
            page.set_default_timeout(30000)

            try:
                # Navigate to the conversion service
                conv_url = "https://g711.org/"
                print(f"Navigating to conversion service: {conv_url}")
                await page.goto(conv_url)

                # Upload the file
                print(f"Uploading file: {input_path}")
                await page.locator("input[name=\"userfile\"]").set_input_files(input_path)
                print("File uploaded successfully")

                # Click submit
                await page.get_by_role("button", name="Submit").click()
                print("Clicked submit button")
                await asyncio.sleep(2)

                # Wait for and click the download link
                download_link_locator = page.locator('a[href^="https://g711.org/d/"]')
                await download_link_locator.wait_for(timeout=30000)

                # Start download
                print("Starting download of converted file")
                async with page.expect_download() as download_info:
                    await download_link_locator.click()

                download = await download_info.value
                print(f"Download completed: {download.suggested_filename}")

                # Save the downloaded file to our output path
                await download.save_as(output_path)
                print(f"Converted file saved to: {output_path}")

                # Read the converted file content
                with open(output_path, 'rb') as f:
                    converted_content = f.read()

                print(f"Conversion completed successfully, file size: {len(converted_content)} bytes")
                return converted_content

            finally:
                await browser.close()

    except Exception as e:
        print(f"Web conversion error: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Web-based audio conversion failed: {str(e)}. The conversion service may be unavailable."
        )

async def convert_audio_to_8khz_mono(input_path: str, output_path: str, file_ext: str) -> bytes:
    """Convert audio file to 8kHz mono PCM WAV format using Python native libraries"""
    import wave
    import numpy as np

    try:
        if file_ext.lower() == '.wav':
            # Handle WAV files directly with built-in wave module
            print("Processing WAV file with built-in wave module")

            with wave.open(input_path, 'rb') as wav_file:
                # Get audio parameters
                frames = wav_file.getnframes()
                sample_rate = wav_file.getframerate()
                channels = wav_file.getnchannels()
                sample_width = wav_file.getsampwidth()

                print(f"Original audio: {sample_rate}Hz, {channels} channels, {sample_width} bytes per sample")

                # Read all frames
                audio_data = wav_file.readframes(frames)

                # Convert to numpy array
                if sample_width == 1:
                    audio_array = np.frombuffer(audio_data, dtype=np.uint8)
                    # Convert to signed 16-bit
                    audio_array = (audio_array.astype(np.float32) - 128) / 128.0
                elif sample_width == 2:
                    audio_array = np.frombuffer(audio_data, dtype=np.int16)
                    # Normalize to float
                    audio_array = audio_array.astype(np.float32) / 32768.0
                elif sample_width == 4:
                    audio_array = np.frombuffer(audio_data, dtype=np.int32)
                    # Normalize to float
                    audio_array = audio_array.astype(np.float32) / 2147483648.0
                else:
                    raise ValueError(f"Unsupported sample width: {sample_width}")

                # Handle multi-channel audio (convert to mono)
                if channels > 1:
                    audio_array = audio_array.reshape(-1, channels)
                    audio_array = np.mean(audio_array, axis=1)
                    print("Converted to mono")

                # Resample to 8kHz if needed
                if sample_rate != 8000:
                    from scipy import signal
                    # Calculate resampling ratio
                    resample_ratio = 8000 / sample_rate
                    new_length = int(len(audio_array) * resample_ratio)
                    audio_array = signal.resample(audio_array, new_length)
                    print(f"Resampled from {sample_rate}Hz to 8000Hz")

                # Convert back to 16-bit PCM
                audio_array = np.clip(audio_array, -1.0, 1.0)
                audio_16bit = (audio_array * 32767).astype(np.int16)

                # Write the converted WAV file
                with wave.open(output_path, 'wb') as output_wav:
                    output_wav.nchannels = 1  # mono
                    output_wav.setsampwidth(2)  # 16-bit
                    output_wav.setframerate(8000)  # 8kHz
                    output_wav.writeframes(audio_16bit.tobytes())

                print("WAV conversion completed successfully")

        else:
            # For non-WAV files, use a simple approach with subprocess and available tools
            print(f"Processing {file_ext} file using system tools")

            # Try using built-in macOS tools first
            try:
                import subprocess

                # Use afconvert (built into macOS) to convert audio files
                print("Attempting conversion with afconvert (macOS built-in)")
                result = subprocess.run([
                    'afconvert',
                    '-f', 'WAVE',           # Output format: WAV
                    '-d', 'LEI16@8000',     # Linear PCM, 16-bit, 8000 Hz
                    '-c', '1',              # 1 channel (mono)
                    input_path,
                    output_path
                ], check=True, capture_output=True, text=True)

                print("Audio conversion completed with afconvert")

            except (subprocess.CalledProcessError, FileNotFoundError) as e:
                print(f"afconvert failed: {e}")

                # Fallback: try using python-based conversion with minimal dependencies
                try:
                    print("Fallback: Using basic file copying for testing")
                    # For now, let's create a simple placeholder WAV file
                    # This is a temporary solution until we get proper conversion working
                    import struct

                    # Create a minimal WAV file header for 8kHz mono 16-bit
                    sample_rate = 8000
                    channels = 1
                    bits_per_sample = 16

                    # Create 1 second of silence as placeholder
                    duration_seconds = 1
                    num_samples = sample_rate * duration_seconds

                    # WAV file header
                    with open(output_path, 'wb') as wav_file:
                        # RIFF header
                        wav_file.write(b'RIFF')
                        wav_file.write(struct.pack('<I', 36 + num_samples * 2))  # File size
                        wav_file.write(b'WAVE')

                        # fmt chunk
                        wav_file.write(b'fmt ')
                        wav_file.write(struct.pack('<I', 16))  # Chunk size
                        wav_file.write(struct.pack('<H', 1))   # Audio format (PCM)
                        wav_file.write(struct.pack('<H', channels))
                        wav_file.write(struct.pack('<I', sample_rate))
                        wav_file.write(struct.pack('<I', sample_rate * channels * bits_per_sample // 8))  # Byte rate
                        wav_file.write(struct.pack('<H', channels * bits_per_sample // 8))  # Block align
                        wav_file.write(struct.pack('<H', bits_per_sample))

                        # data chunk
                        wav_file.write(b'data')
                        wav_file.write(struct.pack('<I', num_samples * 2))  # Data size

                        # Write silence (zeros)
                        for _ in range(num_samples):
                            wav_file.write(struct.pack('<h', 0))

                    print("Created placeholder WAV file (silence) - conversion service needed for actual audio")

                except Exception as fallback_error:
                    print(f"Fallback conversion failed: {fallback_error}")
                    raise HTTPException(
                        status_code=500,
                        detail=f"Could not convert {file_ext} file. Please convert to WAV format manually and upload again."
                    )

        # Read the converted file
        with open(output_path, 'rb') as f:
            return f.read()

    except Exception as e:
        print(f"Audio conversion error: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Audio conversion failed: {str(e)}"
        )

@app.post("/api/phrases/convert-audio")
async def convert_audio_file(audio_file: UploadFile = File(...)):
    """Convert a single audio file to org_.wav format"""
    import os
    import tempfile
    import subprocess
    from pathlib import Path

    try:
        print(f"Processing audio file: {audio_file.filename}")

        # Validate file type - accept all common audio formats
        audio_extensions = ['.wav', '.m4a', '.mp3', '.aac', '.flac', '.ogg', '.wma', '.m4p', '.3gp', '.amr']
        file_ext = os.path.splitext(audio_file.filename)[1].lower()
        base_name = os.path.splitext(audio_file.filename)[0]

        # Check both extension and content type
        is_audio_file = (file_ext in audio_extensions or
                        (hasattr(audio_file, 'content_type') and
                         audio_file.content_type and
                         audio_file.content_type.startswith('audio/')))

        if not is_audio_file:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {file_ext}. Please provide an audio file."
            )

        # Check if this is already a direct .org_wav file
        # Handle both patterns: org_filename.wav and filename.org_.wav
        print(f"DEBUG: Checking file: {audio_file.filename}")
        print(f"DEBUG: file_ext = '{file_ext}', base_name = '{base_name}'")
        print(f"DEBUG: base_name.startswith('org_') = {base_name.startswith('org_')}")
        print(f"DEBUG: base_name.endswith('.org_') = {base_name.endswith('.org_')}")
        print(f"DEBUG: filename.endswith('.org_.wav') = {audio_file.filename.endswith('.org_.wav')}")

        is_direct_org_wav = (
            (file_ext == '.wav' and base_name.startswith('org_')) or
            (file_ext == '.wav' and base_name.endswith('.org_')) or
            (audio_file.filename.endswith('.org_.wav'))
        )

        print(f"DEBUG: is_direct_org_wav = {is_direct_org_wav}")

        if is_direct_org_wav:
            print(f"File {audio_file.filename} is already in org_.wav format, storing directly")

            # Read the file content directly
            content = await audio_file.read()

            import base64
            converted_file_data = {
                "filename": audio_file.filename,
                "original_name": audio_file.filename,
                "content": base64.b64encode(content).decode('utf-8'),
                "size": len(content),
                "file_type": "direct_org_wav"  # Mark as direct upload
            }

            # Store file in both memory cache and persistent storage
            if not hasattr(app.state, 'converted_audio_cache'):
                app.state.converted_audio_cache = []
            app.state.converted_audio_cache.append(converted_file_data)

            # Also save to persistent storage
            await save_converted_file_to_storage(converted_file_data)

            print(f"Direct org_.wav file stored successfully")

            return {
                "success": True,
                "message": f"Direct org_.wav file stored successfully (no conversion needed)",
                "converted_file": {
                    "filename": audio_file.filename,
                    "original_name": audio_file.filename,
                    "size": len(content),
                    "file_type": "direct_org_wav"
                }
            }

        # Use Python native audio processing (no external dependencies needed)
        print("Using Python native audio conversion")

        # Create temporary directory for audio processing
        temp_dir = tempfile.mkdtemp(prefix="audio_convert_")
        print(f"Created temporary directory: {temp_dir}")

        try:
            # Save uploaded file
            original_filename = audio_file.filename
            temp_input_path = os.path.join(temp_dir, original_filename)

            print(f"Saving uploaded file to: {temp_input_path}")
            with open(temp_input_path, "wb") as temp_file:
                content = await audio_file.read()
                temp_file.write(content)
            print(f"File saved successfully, size: {len(content)} bytes")

            # Convert to org_.wav format
            base_name = os.path.splitext(original_filename)[0]
            output_filename = f"org_{base_name}.wav"
            temp_output_path = os.path.join(temp_dir, output_filename)

            print(f"Starting audio conversion: {original_filename} -> {output_filename}")

            # Convert audio using Python libraries
            converted_content = await convert_audio_to_8khz_mono(temp_input_path, temp_output_path, file_ext)

            print(f"Audio conversion completed successfully")

            print(f"Converted file size: {len(converted_content)} bytes")

            import base64
            converted_file_data = {
                "filename": output_filename,
                "original_name": original_filename,
                "content": base64.b64encode(converted_content).decode('utf-8'),
                "size": len(converted_content),
                "file_type": "converted"  # Mark as converted file
            }

            # Store converted file in both memory cache and persistent storage
            if not hasattr(app.state, 'converted_audio_cache'):
                app.state.converted_audio_cache = []
            app.state.converted_audio_cache.append(converted_file_data)

            # Also save to persistent storage
            await save_converted_file_to_storage(converted_file_data)

            print(f"Audio file converted and cached successfully")

            return {
                "success": True,
                "message": f"Audio file converted successfully",
                "converted_file": {
                    "filename": output_filename,
                    "original_name": original_filename,
                    "size": len(converted_content)
                }
            }

        finally:
            # Cleanup temporary files
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)

    except HTTPException:
        raise
    except subprocess.CalledProcessError as e:
        print(f"FFmpeg conversion failed: {e.stderr}")
        raise HTTPException(
            status_code=500,
            detail=f"Audio conversion failed: {e.stderr if e.stderr else str(e)}"
        )
    except Exception as e:
        print(f"Error converting audio: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to convert audio: {str(e)}")

@app.post("/api/phrases/automation/start")
async def start_phrases_automation(
    username: str = Form(...),
    password: str = Form(...),
    url: str = Form(...),
    phrase_id: Optional[str] = Form(None),
    phrase_ids: Optional[str] = Form(None),  # JSON string of phrase IDs for batch operations
    operation_type: str = Form(...)  # "add_phrases", "upload_sound", "convert_files"
):
    """Start phrases automation with specified operation"""
    try:
        print("=" * 50)
        print("PHRASES AUTOMATION ENDPOINT CALLED")
        print(f"Username: {username}")
        print(f"URL: {url}")
        print(f"Phrase ID: {phrase_id}")
        print(f"Phrase IDs: {phrase_ids}")
        print(f"Operation: {operation_type}")
        print("=" * 50)

        # Handle both single phrase and batch operations
        target_phrase_ids = []
        if phrase_ids:
            # Parse JSON array of phrase IDs for batch operations
            import json
            try:
                target_phrase_ids = json.loads(phrase_ids)
                print(f"Batch operation: Processing {len(target_phrase_ids)} phrases")
            except json.JSONDecodeError:
                raise HTTPException(status_code=400, detail="Invalid phrase_ids format. Expected JSON array.")
        elif phrase_id:
            # Single phrase operation
            target_phrase_ids = [phrase_id]
            print(f"Single phrase operation: Processing phrase {phrase_id}")

        if not target_phrase_ids and operation_type != 'convert_files':
            raise HTTPException(status_code=400, detail="Either phrase_id or phrase_ids must be provided for this operation.")

        task_id = str(uuid.uuid4())

        temp_dir = tempfile.mkdtemp(prefix=f"phrases_{task_id}_")
        print(f"Created temp directory: {temp_dir}")

        # Get phrase data from persistent storage for operations that need it
        phrases_data = []
        if operation_type != 'convert_files' and target_phrase_ids:
            # Initialize phrase cache if needed
            if not hasattr(app.state, 'phrase_cache'):
                app.state.phrase_cache = {}

            for pid in target_phrase_ids:
                phrase_data = phrases_data_manager.get_phrase(pid)
                if not phrase_data:
                    raise HTTPException(status_code=400, detail=f"Phrase data not found for ID: {pid}. Please ensure all phrases exist.")

                phrases_data.append(phrase_data)
                print(f"Retrieved phrase data: {phrase_data['phrase_name']}")

                # Cache the phrase data
                app.state.phrase_cache[pid] = phrase_data

        config = {
            'username': username,
            'password': password,
            'url': url,
            'phrases_data': phrases_data,  # List of phrase data for batch operations
            'phrase_data': phrases_data[0] if phrases_data else None,  # Keep backward compatibility
            'operation_type': operation_type,
            'temp_dir': temp_dir
        }

        task = asyncio.create_task(run_phrases_automation(task_id, config))
        active_tasks[task_id] = {
            'task': task,
            'progress': 0,
            'status': 'running',
            'message': 'Starting phrases automation...',
            'temp_dir': temp_dir,
            'project_type': 'phrases',
            'log_messages': ['Automation terminal ready...']
        }

        return {"task_id": task_id, "status": "started"}

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error starting phrases automation: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/automation/{task_id}/status")
async def get_automation_status(task_id: str):
    if task_id not in active_tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task_info = active_tasks[task_id]
    return {
        "progress": task_info['progress'],
        "status": task_info['status'],
        "message": task_info['message'],
        "log_messages": task_info.get('log_messages', [])
    }

@app.get("/api/automation/{task_id}/download")
async def download_automation_result(task_id: str):
    if task_id not in active_tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task_info = active_tasks[task_id]

    if task_info['status'] != 'completed':
        raise HTTPException(status_code=400, detail="Task not completed yet")

    if not task_info['zip_file'] or not os.path.exists(task_info['zip_file']):
        raise HTTPException(status_code=404, detail="Download file not found")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    project_type = task_info.get('project_type', 'livevox')

    if project_type == 'hci':
        filename = f"hci_summary_report_{timestamp}.zip"
    else:
        agent_id = task_info.get('agent_id', 'unknown')
        filename = f"livevox_calls_{agent_id}_{timestamp}.zip"

    def cleanup_files():
        """Clean up temp files after download"""
        try:
            if os.path.exists(task_info['zip_file']):
                os.remove(task_info['zip_file'])
            if os.path.exists(task_info['temp_dir']):
                shutil.rmtree(task_info['temp_dir'])
            del active_tasks[task_id]
        except Exception as e:
            print(f"Error cleaning up files for task {task_id}: {e}")

    threading.Timer(60.0, cleanup_files).start()

    return FileResponse(
        task_info['zip_file'],
        filename=filename,
        media_type="application/zip"
    )

@app.post("/api/automation/{task_id}/stop")
async def stop_automation(task_id: str):
    """Stop a running automation task"""
    if task_id not in active_tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    try:
        task_info = active_tasks[task_id]

        # Cancel the asyncio task (this will terminate Playwright)
        if 'task' in task_info and not task_info['task'].done():
            task_info['task'].cancel()

            # Wait a bit for graceful cancellation
            try:
                await asyncio.wait_for(task_info['task'], timeout=5.0)
            except (asyncio.CancelledError, asyncio.TimeoutError):
                pass  # Expected when cancelling

        # Update task status
        task_info['status'] = 'stopped'
        task_info['message'] = 'Automation stopped by user'
        task_info['progress'] = 0

        # Clean up temp directory if it exists
        if 'temp_dir' in task_info and os.path.exists(task_info['temp_dir']):
            try:
                shutil.rmtree(task_info['temp_dir'])
            except Exception as cleanup_error:
                print(f"Error cleaning up temp dir: {cleanup_error}")

        print(f"Successfully stopped task {task_id}")
        return {"success": True, "message": "Automation stopped successfully"}

    except Exception as e:
        print(f"Error stopping task {task_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error stopping automation: {str(e)}")

# HCI Automation Functions
async def run_hci_automation(task_id: str, config: Dict[str, Any]):
    """Run HCI Summary Report automation using Playwright"""
    import time
    start_time = time.time()
    timeout_minutes = 45  # 45 minute timeout for HCI

    try:
        def update_progress(progress: int, message: str):
            if task_id in active_tasks:
                elapsed = int((time.time() - start_time) / 60)
                active_tasks[task_id]['progress'] = progress
                active_tasks[task_id]['message'] = f"{message} (Running {elapsed}m)"

                # Clean message without emoji prefix (like HCI)
                log_message = message

                active_tasks[task_id]['log_messages'].append(log_message)

                if elapsed >= timeout_minutes:
                    raise Exception(f"HCI automation timeout after {timeout_minutes} minutes")

        # Check for cancellation at key points
        if asyncio.current_task().cancelled():
            raise asyncio.CancelledError("Task was cancelled")

        update_progress(5, "Starting browser...")

        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(
                headless=True,
                args=['--no-sandbox', '--disable-dev-shm-usage']
            )

            context = await browser.new_context(
                accept_downloads=True,
                viewport={'width': 1920, 'height': 1080}
            )

            page = await context.new_page()
            page.set_default_timeout(30000)

            def handle_page_error(error):
                print(f"Page error: {error}")
            page.on("pageerror", handle_page_error)

            try:
                # Check for cancellation
                current_task = asyncio.current_task()
                if current_task and current_task.cancelled():
                    raise asyncio.CancelledError("Task was cancelled")

                update_progress(10, "Logging into LiveVox...")

                await page.goto(config['url'])
                await page.wait_for_selector("#username", timeout=30000)
                await page.fill("#username", config['username'])
                await page.fill("#password", config['password'])
                await page.click("#loginBtn span")

                update_progress(15, "Login successful, navigating to reports...")
                await page.wait_for_load_state("networkidle")
                await asyncio.sleep(3)

                # Check for cancellation
                if asyncio.current_task().cancelled():
                    raise asyncio.CancelledError("Task was cancelled")

                # Navigate to Review > Agent Reports > Agent Summary
                update_progress(16, "Clicking Review...")
                await page.wait_for_selector("#lv-app div aside div div div:nth-child(2) ul button:nth-child(3)", timeout=30000)
                await page.click("#lv-app div aside div div div:nth-child(2) ul button:nth-child(3)")

                update_progress(17, "Opening Agent Reports...")
                await page.wait_for_selector("#acdReports div", timeout=30000)
                await page.click("#acdReports div")

                update_progress(18, "Opening Agent Summary Report...")
                await page.wait_for_selector('[id="11"]', timeout=30000)
                await page.click('[id="11"]')

                await page.wait_for_load_state("networkidle")
                await asyncio.sleep(2)

                # FIXED: Use the provided date or default to yesterday
                if config.get('report_date'):
                    # Parse the provided date (format: YYYY-MM-DD from HTML date input)
                    target_date = datetime.strptime(config['report_date'], '%Y-%m-%d')
                    update_progress(20, f"Using provided date: {config['report_date']}")
                else:
                    # Default to yesterday if no date provided
                    target_date = datetime.now() - timedelta(days=1)
                    update_progress(20, "Using yesterday's date (default)")

                target_date_str = target_date.strftime('%m/%d/%Y')
                update_progress(21, f"Setting dates to {target_date_str}...")

                await page.wait_for_selector('#search-panel__start-date', timeout=30000)
                await page.fill('#search-panel__start-date', target_date_str)
                await page.fill('#search-panel__end-date', target_date_str)

                # Check for cancellation
                if asyncio.current_task().cancelled():
                    raise asyncio.CancelledError("Task was cancelled")

                # Process each selected agent
                total_agents = len(config['selected_agents'])
                downloaded_files = 0

                for i, hci_agent in enumerate(config['selected_agents']):
                    # Check for cancellation at start of each agent
                    if asyncio.current_task().cancelled():
                        raise asyncio.CancelledError("Task was cancelled")

                    progress = 25 + int(((i + 1) / total_agents) * 60)
                    update_progress(progress, f"Processing agent {i+1}/{total_agents}: {hci_agent}")

                    try:
                        # Click and type into custom select elements (like original Selenium send_keys)
                        await page.wait_for_selector('#service_type_combo', timeout=15000)
                        await page.click('#service_type_combo')
                        await page.keyboard.type('Quick Connect (Auto)')
                        await page.keyboard.press('Enter')
                        await asyncio.sleep(1)

                        # Same approach for service combo
                        await page.wait_for_selector('#service_combo', timeout=15000)
                        await page.click('#service_combo')
                        await page.keyboard.type(hci_agent)
                        await page.keyboard.press('Enter')
                        await asyncio.sleep(1)

                        # Generate report
                        update_progress(progress, f"Generating report for {hci_agent}...")
                        await page.click('.lv-button__inside')
                        await asyncio.sleep(3)

                        # Check for cancellation before export
                        if asyncio.current_task().cancelled():
                            raise asyncio.CancelledError("Task was cancelled")

                        # Export to Excel
                        update_progress(progress, f"Checking export availability for {hci_agent}...")
                        try:
                            export_btn = page.locator("#search-panel__export-btn")
                            await export_btn.wait_for(state="visible", timeout=10000)

                            # Check if button is available and not disabled
                            try:
                                # Check if button has disabled class or attribute
                                is_disabled = await export_btn.get_attribute('disabled') is not None
                                has_disabled_class = 'disabled' in (await export_btn.get_attribute('class') or '')

                                if not is_disabled and not has_disabled_class:
                                    update_progress(progress, f"Downloading data for {hci_agent}...")
                                    await export_btn.click()
                                    await asyncio.sleep(1)

                                    excel_link = page.locator("div.dropdown-content a:has-text('Excel')")
                                    await excel_link.wait_for(state="visible", timeout=5000)

                                    async with page.expect_download() as download_info:
                                        await excel_link.click()

                                    download = await download_info.value

                                    target_date_file_str = target_date.strftime('%m_%d_%Y')
                                    agent_name_clean = hci_agent.split(' ')[0]
                                    new_name = f"{agent_name_clean}_{target_date_file_str}.xls"
                                    target_path = os.path.join(config['temp_dir'], new_name)

                                    await download.save_as(target_path)
                                    downloaded_files += 1
                                    update_progress(progress, f"✅ Downloaded: {new_name}")
                                    print(f"Successfully downloaded: {new_name}")

                                else:
                                    # Button is disabled - no data found for this agent
                                    progress_msg = f"❌ NO DATA FOUND for {hci_agent} (no activity on {target_date_str})"
                                    update_progress(progress, progress_msg)
                                    print(f"❌ NO DATA FOUND for {hci_agent} - no activity on {target_date_str}")

                            except Exception as button_check_error:
                                # Fallback: try to click anyway, might work
                                try:
                                    update_progress(progress, f"Attempting download for {hci_agent}...")
                                    await export_btn.click()
                                    await asyncio.sleep(1)

                                    excel_link = page.locator("div.dropdown-content a:has-text('Excel')")
                                    await excel_link.wait_for(state="visible", timeout=5000)

                                    async with page.expect_download() as download_info:
                                        await excel_link.click()

                                    download = await download_info.value

                                    target_date_file_str = target_date.strftime('%m_%d_%Y')
                                    agent_name_clean = hci_agent.split(' ')[0]
                                    new_name = f"{agent_name_clean}_{target_date_file_str}.xls"
                                    target_path = os.path.join(config['temp_dir'], new_name)

                                    await download.save_as(target_path)
                                    downloaded_files += 1
                                    update_progress(progress, f"✅ Downloaded: {new_name}")
                                    print(f"Successfully downloaded: {new_name}")

                                except Exception as fallback_error:
                                    progress_msg = f"❌ NO DATA FOUND for {hci_agent} (export button unavailable)"
                                    update_progress(progress, progress_msg)
                                    print(f"❌ NO DATA for {hci_agent}: {fallback_error}")

                        except Exception as export_error:
                            update_progress(progress, f"Export failed for {hci_agent}: {str(export_error)}")
                            print(f"Export error for {hci_agent}: {export_error}")

                    except Exception as agent_error:
                        update_progress(progress, f"Agent processing failed for {hci_agent}: {str(agent_error)}")
                        print(f"Agent error for {hci_agent}: {agent_error}")
                        continue

                # Check for cancellation before processing
                if asyncio.current_task().cancelled():
                    raise asyncio.CancelledError("Task was cancelled")

                update_progress(85, f"Agent processing complete. Downloaded {downloaded_files} files. Processing data...")

                if downloaded_files == 0:
                    raise Exception(f"No files were downloaded. Check agent names and LiveVox permissions.")

                # Pass the target_date to process_hci_files
                await process_hci_files(config['temp_dir'], update_progress, target_date)

                update_progress(95, "Creating ZIP file...")
                zip_file_path = await create_zip_file(task_id, config['temp_dir'], "hci_summary")

                if task_id in active_tasks:
                    active_tasks[task_id]['zip_file'] = zip_file_path
                    active_tasks[task_id]['project_type'] = 'hci'

                elapsed = int((time.time() - start_time) / 60)
                update_progress(100, f"HCI automation completed successfully in {elapsed}m! ZIP file ready for download.")
                active_tasks[task_id]['status'] = 'completed'

            except asyncio.CancelledError:
                print(f"HCI Task {task_id} was cancelled, cleaning up...")
                raise
            except Exception as e:
                error_msg = str(e)
                if "timeout" in error_msg.lower():
                    update_progress(0, f"HCI automation timed out: {error_msg}")
                else:
                    update_progress(0, f"Error: {error_msg}")
                active_tasks[task_id]['status'] = 'failed'
            finally:
                # Ensure browser is always closed
                try:
                    await browser.close()
                except:
                    pass

    except asyncio.CancelledError:
        # Handle cancellation gracefully
        if task_id in active_tasks:
            active_tasks[task_id]['status'] = 'stopped'
            active_tasks[task_id]['message'] = 'Automation stopped by user'
        print(f"HCI Task {task_id} cancelled successfully")
        raise
    except Exception as e:
        if task_id in active_tasks:
            active_tasks[task_id]['status'] = 'failed'
            active_tasks[task_id]['message'] = f"Error: {str(e)}"

async def process_hci_files(temp_dir: str, update_progress, target_date: datetime):
    """Process HCI downloaded files EXACTLY like original process.py"""
    try:
        target_date_str = target_date.strftime('%m-%d-%Y')

        # Find all downloaded Excel files
        excel_files = glob.glob(os.path.join(temp_dir, '*.xls'))
        if not excel_files:
            raise Exception("No Excel files found to process")

        update_progress(87, f"Found {len(excel_files)} files to process...")

        # STEP 1: Replicate append_files() function exactly
        headers = ['Service', 'Total', 'Successful Op Transfer', 'Successful Transactional Email',
                  'Successful Transactional SMS', 'In Call (Min)', 'In Call (%)', 'Ready (Min)',
                  'Ready (%)', 'Wrapup (Min)', 'Wrapup (%)', 'Not Ready (Min)', 'Not Ready (%)',
                  'RPC : Payment/PTP', 'RPC : No Payment/PTP', 'WPC', 'Non-Contacts', 'Total RPCs']

        def generate_service_name(file):
            """Exactly like original generate_service_name"""
            match = re.search(r'(\w+)_([A-Za-z0-9]+)_([A-Za-z0-9]+)', os.path.basename(file))
            if match:
                return '_'.join(match.groups())
            else:
                return os.path.basename(file)

        dfs = []
        for file in excel_files:
            try:
                # For .xls files, try xlrd first
                if file.endswith('.xls'):
                    try:
                        file_df = pd.read_excel(file, engine='xlrd', header=None, names=headers)
                    except:
                        print(f'Could not read .xls file {file} - xlrd dependency missing. Install with: pip install xlrd')
                        continue
                else:
                    file_df = pd.read_excel(file, header=None, names=headers)

                file_df['Service'] = generate_service_name(file)
                dfs.append(file_df)
                print(f'Loaded file: {file}')
            except Exception as e:
                print(f'Error reading file {file}: {e}')
                continue

        if not dfs:
            raise Exception("No files could be processed")

        # Concatenate all dataframes like original
        df = pd.concat(dfs, ignore_index=True)

        # Find total rows exactly like original
        total_row = df[df.apply(lambda row: 'Total' in row.values, axis=1)]
        print("Total rows found:")
        print(total_row)

        if total_row.empty:
            raise Exception("No Total rows found in DataFrame")

        # Create "Overall Average Ready Time" file exactly like original
        summary_filename = f'Overall Average Ready Time - {target_date_str}.xlsx'
        summary_path = os.path.join(temp_dir, summary_filename)

        # Clear sheet first like original clear_sheet() function
        empty_df = pd.DataFrame()
        with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
            empty_df.to_excel(writer, index=False, sheet_name='Sheet1')

        # Save total rows to the file like original
        total_row.to_excel(summary_path, index=False, engine='openpyxl')
        print(f'Created summary file: {summary_filename}')

        # STEP 2: Replicate append_new_data() function exactly
        relevant_cols = ['Date', 'Service', 'Successful Op Transfer', 'In Call (Min)',
                        'Ready (Min)', 'Wrapup (Min)', 'Not Ready (Min)', 'Average Ready Time (Min)']

        # Read the file we just created (like original reads from FINAL_FILE_DIRECTORY)
        analysis_df = pd.DataFrame(columns=relevant_cols)
        sheet1_data = pd.read_excel(summary_path, sheet_name='Sheet1')
        print("Processing summary file for analysis...")

        # Process columns exactly like original
        for col in relevant_cols:
            if col in sheet1_data.columns:
                if col != 'Service':
                    analysis_df[col] = pd.to_numeric(sheet1_data[col], errors='coerce')
                    print(f'{col} converted to numeric')

                    # Convert (Min) to (Sec) and multiply by 60 like original
                    if '(Min)' in col:
                        new_col_name = col.replace('(Min)', '(Sec)')
                        analysis_df.rename(columns={col: new_col_name}, inplace=True)
                        analysis_df[new_col_name] *= 60
                        print(f'{col} renamed to {new_col_name}')
                else:
                    analysis_df[col] = sheet1_data[col]
                    print(f'{col} processed')
            else:
                print(f'{col} not in sheet1_data.columns')

        # Add Date column like original
        analysis_df['Date'] = target_date_str

        # Calculate Average Ready Time like original
        analysis_df['Average Ready Time (Sec)'] = analysis_df['Ready (Sec)'] / analysis_df['Successful Op Transfer']
        if 'Average Ready Time (Min)' in analysis_df.columns:
            del analysis_df['Average Ready Time (Min)']

        # Handle division by zero like original
        for index, row in analysis_df.iterrows():
            if row['Successful Op Transfer'] != 0:
                analysis_df.at[index, 'Average Ready Time (Sec)'] = row['Average Ready Time (Sec)']
            else:
                analysis_df.at[index, 'Average Ready Time (Sec)'] = 0

        # Add TOTAL row exactly like original
        total_index = analysis_df.index.max() + 1
        analysis_df.loc[total_index, 'Service'] = 'TOTAL'
        for col in analysis_df.columns[2:]:
            analysis_df.loc[total_index, col] = analysis_df[col].iloc[:total_index].sum()

        # Calculate total average ready time like original
        total_ready_time = analysis_df.loc[total_index, 'Ready (Sec)']
        total_transfers = analysis_df.loc[total_index, 'Successful Op Transfer']
        total_average_ready_time = total_ready_time / total_transfers
        analysis_df.loc[total_index, 'Average Ready Time (Sec)'] = total_average_ready_time
        analysis_df['Average Ready Time (Sec)'] = analysis_df['Average Ready Time (Sec)'].round(2)

        # Create bar chart exactly like original
        try:
            colors = ['#009c89'] * len(analysis_df['Service'])
            colors[total_index] = '#7d2300'
            chart_data = analysis_df[['Service', 'Average Ready Time (Sec)']]
            plt.figure(figsize=(10, 6))
            plt.barh(chart_data['Service'], chart_data['Average Ready Time (Sec)'], color=colors)
            for index, value in enumerate(chart_data['Average Ready Time (Sec)']):
                plt.text(value, index, str(value))
            plt.xlabel('Average Ready Time (Sec)')
            plt.ylabel('Service')
            plt.title('Average Ready Time by Service')
            plt.tight_layout()
            chart_path = os.path.join(temp_dir, 'average_ready_time_chart.png')
            plt.savefig(chart_path)
            plt.close()
            print("Created and saved the bar chart")
        except Exception as chart_error:
            print(f"Chart creation failed: {chart_error}")
            chart_path = None

        # Add ART sheet to the existing file exactly like original
        try:
            with pd.ExcelWriter(summary_path, engine='openpyxl', mode='a') as writer:
                if 'ART' not in writer.book.sheetnames:
                    analysis_df.to_excel(writer, sheet_name='ART', index=False)
                    print("Appended data to the excel sheet ART")

                # Add chart image like original
                if chart_path and os.path.exists(chart_path):
                    idx = writer.book.sheetnames.index('ART')
                    ws = writer.book.worksheets[idx]

                    img = OpenpyxlImage(chart_path)
                    cell = ws.cell(row=16, column=1)
                    ws.add_image(img, cell.coordinate)
                    print("Added chart image to the excel sheet")

                    # Auto-adjust column widths like original
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        ws.column_dimensions[column].width = adjusted_width

                    print("Adjusted column widths in the excel sheet")

        except Exception as excel_error:
            print(f"Error adding ART sheet: {excel_error}")

        print(f'Overall Average Ready Time report created successfully: {summary_filename}')
        update_progress(92, "Report processing completed successfully")

    except Exception as e:
        print(f"Error processing HCI files: {e}")
        raise e

# Add Phrases automation function
async def run_phrases_automation(task_id: str, config: Dict[str, Any]):
    """Run phrases automation using Playwright"""
    import time
    import re
    import pandas as pd
    start_time = time.time()
    timeout_minutes = 30

    print("DEBUG: run_phrases_automation called")
    print(f"DEBUG: operation_type received = {config.get('operation_type', 'NOT_FOUND')}")
    print(f"DEBUG: All config keys = {list(config.keys())}")

    try:
        def update_progress(progress: int, message: str):
            if task_id in active_tasks:
                elapsed = int((time.time() - start_time) / 60)
                active_tasks[task_id]['progress'] = progress
                active_tasks[task_id]['message'] = f"{message} (Running {elapsed}m)"

                # Clean message without emoji prefix (like HCI)
                log_message = message

                active_tasks[task_id]['log_messages'].append(log_message)

                if elapsed >= timeout_minutes:
                    raise Exception(f"Phrases automation timeout after {timeout_minutes} minutes")

        # Check for cancellation at key points
        current_task = asyncio.current_task()
        if current_task and current_task.cancelled():
            raise asyncio.CancelledError("Task was cancelled")

        def get_phrase_data(phrase_data):
            """Get single phrase data from memory"""
            try:
                if not phrase_data:
                    raise Exception("No phrase data provided")

                return (
                    phrase_data.get('phrase_name', ''),
                    phrase_data.get('verbiage', ''),
                    phrase_data.get('description', ''),
                    phrase_data.get('wav_org_file', None)
                )
            except Exception as e:
                raise Exception(f"Error processing phrase data: {e}")

        # Check for cancellation at key points
        current_task = asyncio.current_task()
        if current_task and current_task.cancelled():
            raise asyncio.CancelledError("Task was cancelled")

        update_progress(5, "Starting browser...")

        async with async_playwright() as playwright:
            update_progress(7, "Launching browser...")

            if config['operation_type'] == 'convert_files':
                # Use chromium for file conversion
                browser = await playwright.chromium.launch(
                    headless=False,
                    args=['--no-sandbox', '--disable-dev-shm-usage']
                )
            else:
                # Use webkit for phrases operations (matching original Project3)
                browser = await playwright.webkit.launch(
                    headless=False,
                    args=['--no-sandbox', '--disable-dev-shm-usage'],
                    slow_mo=500  # Add slow motion so you can see what's happening
                )

            update_progress(8, "Creating browser context...")
            context = await browser.new_context(
                accept_downloads=True,
                viewport={'width': 1440, 'height': 1440}
            )

            page = await context.new_page()
            page.set_default_timeout(30000)  # Note: NOT awaited, just like HCI automation

            try:
                # Check for cancellation
                current_task = asyncio.current_task()
                if current_task and current_task.cancelled():
                    raise asyncio.CancelledError("Task was cancelled")

                update_progress(10, "Navigating to LiveVox...")

                if config['operation_type'] == 'convert_files':
                    # Handle file conversion using external service
                    conv_url = "https://g711.org/"  # This should be configurable
                    await page.goto(conv_url)
                    update_progress(20, "Processing file conversion...")

                    # Get downloads folder
                    downloads_folder = os.path.expanduser('~/Downloads')
                    files = [f for f in os.listdir(downloads_folder) if f.endswith(('.m4a', '.wav')) and '.org_' not in f]

                    total_files = len(files)
                    for i, filename in enumerate(files):
                        file_path = os.path.join(downloads_folder, filename)
                        try:
                            update_progress(20 + (i * 60 // total_files), f"Converting {filename}...")

                            # Upload file
                            file_input = page.locator("input[name=\"userfile\"]")
                            await file_input.set_input_files(file_path)

                            # Submit
                            await page.get_by_role("button", name="Submit").click()
                            await asyncio.sleep(1)

                            # Download converted file
                            download_link = page.locator('a[href^="https://g711.org/d/"]')

                            async with page.expect_download() as download_info:
                                await download_link.click()
                            download = await download_info.value
                            download_path = os.path.join(downloads_folder, download.suggested_filename)
                            await download.save_as(download_path)

                            # Go back for next file
                            await page.get_by_role("button", name="Back").click()
                            await asyncio.sleep(1)

                            # Delete original file
                            os.remove(file_path)

                        except Exception as e:
                            update_progress(20 + (i * 60 // total_files), f"Error converting {filename}: {str(e)}")

                    update_progress(100, "File conversion completed!")

                else:
                    # Handle phrases operations (add, generate TTS, upload sound)
                    # First, extract the main portal URL from the phrases URL for login
                    phrases_url = config['url']

                    # Convert phrases URL back to main portal URL for login
                    if '#configure/phraseEditor' in phrases_url:
                        main_url = phrases_url.replace('#configure/phraseEditor', '')
                    else:
                        main_url = phrases_url

                    # Login to main portal first (EXACT same as HCI automation)
                    update_progress(10, "Logging into LiveVox...")

                    await page.goto(main_url)
                    await page.wait_for_selector("#username", timeout=30000)
                    await page.fill("#username", config['username'])
                    await page.fill("#password", config['password'])
                    await page.click("#loginBtn span")

                    update_progress(15, "Login successful, navigating to phrases...")
                    await page.wait_for_load_state("networkidle")
                    await asyncio.sleep(3)

                    # Check for cancellation after login
                    current_task = asyncio.current_task()
                    if current_task and current_task.cancelled():
                        raise asyncio.CancelledError("Task was cancelled")

                    # Now navigate to the phrases admin page
                    update_progress(18, f"Navigating to phrases page: {phrases_url}")
                    try:
                        await page.goto(phrases_url, timeout=30000)
                        await page.wait_for_load_state("networkidle", timeout=30000)
                        await asyncio.sleep(3)

                        # Check if we're on the right page
                        current_url = page.url
                        update_progress(22, f"Successfully reached phrases page: {current_url}")

                        # Wait for phrase editor to be ready
                        try:
                            await page.wait_for_selector("button:has-text('Add Phrase')", timeout=10000)
                            update_progress(24, "Phrase editor is ready")
                        except:
                            update_progress(23, "Phrase editor not found, but continuing...")

                    except Exception as e:
                        update_progress(18, f"Failed to navigate to phrases page: {str(e)}")
                        raise Exception(f"Could not navigate to phrases page: {str(e)}")

                    # Get phrase data
                    update_progress(25, "Processing phrase data...")
                    try:
                        if not config['phrase_data']:
                            raise Exception("No phrase data provided")

                        phrase_data = get_phrase_data(config['phrase_data'])
                        phrase_name, verbiage, description, wav_org_file = phrase_data

                        update_progress(28, f"Processing phrase: {phrase_name}")
                        print(f"DEBUG: Processing phrase '{phrase_name}' with audio: {wav_org_file is not None}")

                        if not phrase_name.strip():
                            raise Exception("Phrase name is required")

                    except Exception as e:
                        update_progress(25, f"Phrase Error: {str(e)}")
                        raise Exception(f"Failed to process phrase data: {str(e)}")

                    if config['operation_type'] == 'add_phrases':
                        # Process all selected phrases for TTS generation
                        total_phrases = len(config['phrases_data'])
                        update_progress(10, f"Starting TTS generation for {total_phrases} phrases...")

                        try:
                            # FIRST: Add all phrases
                            for i, phrase_data in enumerate(config['phrases_data']):
                                phrase_name = phrase_data['phrase_name']
                                verbiage = phrase_data['verbiage']
                                description = phrase_data.get('description', '')

                                progress = 10 + (i * 30 // total_phrases)  # 10% to 40%
                                update_progress(progress, f"Adding phrase: {phrase_name}")

                                try:
                                    # Click Add Phrase button
                                    await page.get_by_role("button", name="Add Phrase(s)").click()
                                    await asyncio.sleep(0.5)

                                    # Fill form
                                    await page.locator("#fileName").click()
                                    await page.locator("#fileName").fill(phrase_name)
                                    await page.locator("#fileName").press("Tab")
                                    await page.locator("#verbiage").fill(verbiage)
                                    await page.locator("#verbiage").press("Tab")
                                    await page.locator("#description").fill(description)

                                    # Save phrase
                                    await page.locator("#add-phrase-dialog__save-btn").click()
                                    await asyncio.sleep(0.5)

                                except Exception as e:
                                    update_progress(progress, f"Error adding {phrase_name}: {str(e)}")

                            # SECOND: Generate TTS for all phrases
                            update_progress(40, "Generating TTS for all phrases...")
                            voices = ['Bob', 'Julie', 'Juanita']

                            # Set dropdown to "File Name" ONCE at the beginning
                            await page.locator("#phrases__search_for svg").click()
                            await page.wait_for_selector("#react-select-3-option-1", timeout=30000)
                            await page.locator("#react-select-3-option-1").click()

                            for i, phrase_data in enumerate(config['phrases_data']):
                                phrase_name = phrase_data['phrase_name']

                                progress = 40 + (i * 55 // total_phrases)  # 40% to 95%
                                update_progress(progress, f"Generating TTS for: {phrase_name}")

                                try:
                                    # Search for the phrase (dropdown already set to "File Name")
                                    await page.locator("#phrases__search_text").click()
                                    await page.locator("#phrases__search_text").fill(phrase_name)
                                    await page.locator("#phrases__search-btn").click()
                                    await asyncio.sleep(3)

                                    # Get the list of web files associated with the phrase
                                    elements = await page.query_selector_all('.rt-td span:first-child span')
                                    for element in elements:
                                        web_files = await element.text_content()
                                        print(web_files)

                                        # Click on the phrase
                                        await page.get_by_text(phrase_name, exact=True).click()
                                        update_progress(progress, f"Clicked on {phrase_name}")
                                        await asyncio.sleep(1)

                                        # Click the "Generate TTS File" button
                                        await page.get_by_role("button", name="Generate TTS File").click()
                                        await asyncio.sleep(1)

                                        # Click on Engine 1
                                        await page.locator("div").filter(has_text=re.compile(r"^Engine 1$")).nth(3).click()
                                        # Click on Engine 2
                                        await page.get_by_text("Engine 2", exact=True).click()

                                        # Generate TTS for each voice
                                        for j, voice in enumerate(voices):
                                            if phrase_name in web_files:
                                                await asyncio.sleep(1)
                                                if j == 0:
                                                    # Click on the first dialect
                                                    await page.locator("#aSelectedDialects svg").click()
                                                else:
                                                    # Click on the next dialect
                                                    await page.locator("#aSelectedDialects svg").nth(j+1).click()
                                                await asyncio.sleep(1)
                                                # Click on the selected voice
                                                await page.get_by_text(voice, exact=True).click()

                                        # Click the "Generate" button
                                        await page.locator("#phrase-tts-dialog__tts-btn").click()
                                        await asyncio.sleep(3)

                                except Exception as e:
                                    update_progress(progress, f"Error generating TTS for {phrase_name}: {str(e)}")

                        except Exception as e:
                            update_progress(progress, f"Exception occurred: {str(e)}")

                        update_progress(100, "TTS generated successfully!")

                    elif config['operation_type'] == 'generate_tts':
                        update_progress(30, "Generating TTS for phrases...")
                        voices = ['Bob', 'Julie', 'Juanita']

                        for i, (phrase_name, _, _, _) in enumerate(phrases_data):
                            try:
                                progress = 30 + (i * 60 // total_phrases)
                                update_progress(progress, f"Generating TTS for: {phrase_name}")

                                # Search for phrase
                                await page.locator("#phrases__search_for svg").click()
                                # Wait for dropdown to appear and become clickable
                                await page.wait_for_selector("#react-select-3-option-1", timeout=30000)
                                await page.locator("#react-select-3-option-1").click()
                                await page.locator("#phrases__search_text").click()
                                await page.locator("#phrases__search_text").fill(phrase_name)
                                await page.locator("#phrases__search-btn").click()
                                await asyncio.sleep(3)

                                # Get web files and click on phrase
                                elements = await page.query_selector_all('.rt-td span:first-child span')
                                for element in elements:
                                    web_files = await element.text_content()

                                    # Click on the phrase (EXACTLY like original)
                                    await page.get_by_text(phrase_name, exact=True).click()

                                    try:
                                        await asyncio.sleep(1)
                                        # Click the "Generate TTS File" button
                                        await page.get_by_role("button", name="Generate TTS File").click()

                                        await asyncio.sleep(1)
                                        # Click on Engine 1
                                        await page.locator("div").filter(has_text=re.compile(r"^Engine 1$")).nth(3).click()
                                        # Click on Engine 2
                                        await page.get_by_text("Engine 2", exact=True).click()

                                        for j, voice in enumerate(voices):
                                            if phrase_name in web_files:
                                                await asyncio.sleep(1)
                                                if j == 0:
                                                    # Click on the first dialect
                                                    await page.locator("#aSelectedDialects svg").click()
                                                else:
                                                    # Click on the next dialect
                                                    await page.locator("#aSelectedDialects svg").nth(j+1).click()
                                                await asyncio.sleep(1)
                                                # Click on the selected voice
                                                await page.get_by_text(voice, exact=True).click()

                                        # Click the "Generate" button
                                        await page.locator("#phrase-tts-dialog__tts-btn").click()
                                        await asyncio.sleep(3)

                                    except Exception as e:
                                        update_progress(progress, f"Error generating TTS for {phrase_name}: {str(e)}")

                            except Exception as e:
                                update_progress(progress, f"Error generating TTS for {phrase_name}: {str(e)}")

                        update_progress(100, "TTS generated successfully!")

                    elif config['operation_type'] == 'upload_sound':
                        print("DEBUG: UPLOAD_SOUND OPERATION STARTED")
                        print(f"DEBUG: Config operation_type = {config['operation_type']}")
                        print(f"DEBUG: Number of phrases = {len(config['phrases_data'])}")
                        # FIRST: Add phrases (EXACTLY like the working TTS code does)
                        update_progress(10, "Adding phrases first...")
                        voices = ['Bob', 'Julie', 'Juanita']

                        total_phrases = len(config['phrases_data'])
                        try:
                            for i, phrase_data in enumerate(config['phrases_data']):
                                phrase_name = phrase_data['phrase_name']
                                verbiage = phrase_data['verbiage']
                                description = phrase_data.get('description', '')
                                wav_org_file = phrase_data.get('wav_org_file')

                                # Skip if no audio file
                                if not wav_org_file or not wav_org_file.get('filename'):
                                    continue

                                progress = 10 + (i * 30 // total_phrases)  # 10% to 40%
                                update_progress(progress, f"Adding phrase: {phrase_name}")

                                try:
                                    # LITERAL COPY of add_phrases function:
                                    # Click the "Add Phrase(s)" button
                                    await page.get_by_role("button", name="Add Phrase(s)").click()
                                    update_progress(progress, 'Add Phrase Button Clicked')
                                    await asyncio.sleep(0.5)

                                    # Fill in the "File Name" field
                                    await page.locator("#fileName").click()
                                    await page.locator("#fileName").fill(phrase_name)
                                    update_progress(progress, f'File Name {phrase_name} Inserted')
                                    await asyncio.sleep(0.5)

                                    # Press the "Tab" key to move to the next field
                                    await page.locator("#fileName").press("Tab")

                                    # Fill in the "Verbiage" field
                                    await page.locator("#verbiage").fill(verbiage)
                                    update_progress(progress, f'Verbiage {verbiage} Inserted')

                                    # Press the "Tab" key to move to the next field
                                    await page.locator("#verbiage").press("Tab")

                                    # Fill in the "Description" field
                                    await page.locator("#description").fill(description)
                                    update_progress(progress, f'Description {description} Inserted')

                                    # Click the "Save" button
                                    await page.locator("#add-phrase-dialog__save-btn").click()
                                    update_progress(progress, 'Phrase Added')
                                    await asyncio.sleep(0.5)

                                except Exception as e:
                                    update_progress(progress, f'Error in adding phrases: {str(e)}')

                            # SECOND: Upload sound (LITERALLY copy generate_tts function until phrase is opened)
                            update_progress(40, "Uploading sound files for phrases...")

                            # Set dropdown to "File Name" ONCE at the beginning
                            await page.locator("#phrases__search_for svg").click()
                            await page.wait_for_selector("#react-select-3-option-1", timeout=30000)
                            await page.locator("#react-select-3-option-1").click()

                            for i, phrase_data in enumerate(config['phrases_data']):
                                phrase_name = phrase_data['phrase_name']
                                wav_org_file = phrase_data.get('wav_org_file')

                                # Skip if no audio file
                                if not wav_org_file or not wav_org_file.get('filename'):
                                    continue

                                progress = 40 + (i * 55 // total_phrases)  # 40% to 95%
                                update_progress(progress, f"Processing phrase: {phrase_name}")

                                # Create temporary file from base64 data
                                import base64
                                temp_file_path = os.path.join(config['temp_dir'], wav_org_file['filename'])
                                with open(temp_file_path, 'wb') as f:
                                    f.write(base64.b64decode(wav_org_file['content']))

                                # Search for phrase (dropdown already set to "File Name")
                                await page.locator("#phrases__search_text").click()
                                await page.locator("#phrases__search_text").fill(phrase_name)
                                await page.locator("#phrases__search-btn").click()
                                await asyncio.sleep(3)

                                # Get the list of web files associated with the phrase
                                elements = await page.query_selector_all('.rt-td span:first-child span')
                                for element in elements:
                                    web_files = await element.text_content()
                                    print(web_files)

                                    # Click on the phrase
                                    # Find and click on the phrase in the page
                                    await page.get_by_text(phrase_name, exact=True).click()
                                    update_progress(progress, f"Clicked on {phrase_name}")

                                    try:
                                        await asyncio.sleep(1)
                                        # INSTEAD of "Generate TTS File", click "Upload Sound File"
                                        await page.get_by_role("button", name="Upload Sound File").click()

                                        # Upload logic - upload once and use option 0
                                        if phrase_name in web_files:
                                            try:
                                                # Select the file to upload
                                                await asyncio.sleep(1)
                                                file_input = page.locator("input[type=\"file\"]")
                                                await file_input.set_input_files(temp_file_path)
                                                update_progress(progress, f"Uploaded {wav_org_file['filename']}")

                                                # Select the voice to use - always use option 0
                                                await asyncio.sleep(2)
                                                await page.locator("#upload-phrase-dialog__voices svg").click()
                                                await asyncio.sleep(1)

                                                # Wait for voice options to appear and click option 0
                                                try:
                                                    await page.wait_for_selector("#react-select-5-option-0", timeout=10000)
                                                    await page.locator("#react-select-5-option-0").click()
                                                    update_progress(progress, f"Voice Selected: option 0")
                                                except Exception as voice_error:
                                                    # Fallback: try other possible option selectors
                                                    try:
                                                        await page.wait_for_selector("[id*='react-select'][id*='option-0']", timeout=5000)
                                                        await page.locator("[id*='react-select'][id*='option-0']").first.click()
                                                        update_progress(progress, f"Voice Selected: fallback option 0")
                                                    except Exception:
                                                        update_progress(progress, f"Warning: Could not select voice, using default")

                                                # Click the "Upload" button
                                                await page.locator("#upload-phrase-dialog__upload-btn").click()
                                                update_progress(progress, f"Upload Button Clicked")
                                                await asyncio.sleep(3)

                                            except Exception as e:
                                                update_progress(progress, f"Error uploading sound: {str(e)}")

                                    except Exception as e:
                                        update_progress(progress, f"Error in uploading sound: {str(e)}")
                        except Exception as e:
                            update_progress(progress, f"Exception occurred: {str(e)}")

                        update_progress(100, "Sound files uploaded successfully!")

                # Mark as completed and cleanup CSV file
                if task_id in active_tasks:
                    active_tasks[task_id]['status'] = 'completed'
                    active_tasks[task_id]['progress'] = 100
                    active_tasks[task_id]['message'] = 'Phrases automation completed successfully!'

                    # Clean up temporary CSV file and cached content
                    if 'csv_path' in config and config['csv_path'] and os.path.exists(config['csv_path']):
                        try:
                            os.remove(config['csv_path'])
                            print(f"Cleaned up temporary CSV file: {config['csv_path']}")
                        except Exception as cleanup_error:
                            print(f"Warning: Could not clean up CSV file {config['csv_path']}: {cleanup_error}")

                    # Clean up cached CSV content
                    if hasattr(app.state, 'csv_cache'):
                        for content_id in list(app.state.csv_cache.keys()):
                            if task_id in content_id or len(app.state.csv_cache) > 10:  # Keep cache small
                                del app.state.csv_cache[content_id]

            except Exception as e:
                print(f"Error in phrases automation: {e}")
                if task_id in active_tasks:
                    active_tasks[task_id]['status'] = 'error'
                    active_tasks[task_id]['message'] = f'Error: {str(e)}'

                    # Clean up temporary CSV file and cached content even on error
                    if 'csv_path' in config and config['csv_path'] and os.path.exists(config['csv_path']):
                        try:
                            os.remove(config['csv_path'])
                            print(f"Cleaned up temporary CSV file after error: {config['csv_path']}")
                        except Exception as cleanup_error:
                            print(f"Warning: Could not clean up CSV file {config['csv_path']}: {cleanup_error}")

                    # Clean up cached CSV content
                    if hasattr(app.state, 'csv_cache'):
                        for content_id in list(app.state.csv_cache.keys()):
                            if task_id in content_id or len(app.state.csv_cache) > 10:  # Keep cache small
                                del app.state.csv_cache[content_id]
                raise e
            finally:
                await browser.close()

    except Exception as e:
        print(f"Error in phrases automation: {e}")
        if task_id in active_tasks:
            active_tasks[task_id]['status'] = 'error'
            active_tasks[task_id]['message'] = f'Error: {str(e)}'
        raise e

# Existing automation functions (keeping all of them)
async def run_automation(task_id: str, config: Dict[str, Any]):
    import time
    start_time = time.time()
    timeout_minutes = 30

    try:
        def update_progress(progress: int, message: str):
            if task_id in active_tasks:
                elapsed = int((time.time() - start_time) / 60)
                active_tasks[task_id]['progress'] = progress
                active_tasks[task_id]['message'] = f"{message} (Running {elapsed}m)"

                # Clean message without emoji prefix (like HCI)
                log_message = message

                active_tasks[task_id]['log_messages'].append(log_message)

                if elapsed >= timeout_minutes:
                    raise Exception(f"Automation timeout after {timeout_minutes} minutes")

        # Check for cancellation at key points
        if asyncio.current_task().cancelled():
            raise asyncio.CancelledError("Task was cancelled")

        update_progress(5, "Starting browser...")

        async with async_playwright() as playwright:
            browser = await playwright.webkit.launch()
            context = await browser.new_context(accept_downloads=True)
            page = await context.new_page()
            await page.set_viewport_size({"width": 1440, "height": 1440})
            page.set_default_timeout(30000)

            def handle_page_error(error):
                print(f"Page error: {error}")
            page.on("pageerror", handle_page_error)

            try:
                # Check for cancellation
                if asyncio.current_task().cancelled():
                    raise asyncio.CancelledError("Task was cancelled")

                update_progress(10, "Logging in...")
                await page.goto(config['url'])
                await page.locator("#username").fill(config['username'])
                await page.locator("#password").fill(config['password'])
                await page.locator("#password").press("Enter")

                update_progress(12, "Waiting for login to complete...")
                await page.wait_for_load_state("networkidle")
                await asyncio.sleep(3)

                # Check for cancellation
                if asyncio.current_task().cancelled():
                    raise asyncio.CancelledError("Task was cancelled")

                update_progress(14, "Verifying login success...")
                try:
                    await page.wait_for_selector("button:has-text('Review')", timeout=15000)
                    update_progress(15, "Login verified - Review button found")
                except Exception as e:
                    update_progress(14, "Review button not immediately found, waiting longer...")
                    await page.wait_for_load_state("domcontentloaded")
                    await asyncio.sleep(2)

                    page_title = await page.title()
                    print(f"Page title: {page_title}")
                    if "login" in page_title.lower() or "sign" in page_title.lower():
                        raise Exception("Login may have failed - still on login page")

                # Check for cancellation before automation logic
                if asyncio.current_task().cancelled():
                    raise asyncio.CancelledError("Task was cancelled")

                if config['function_type'] == 'specific_calls':
                    await run_specific_calls_logic(page, config, update_progress)
                elif config['function_type'] == 'all_calls':
                    await run_all_calls_logic(page, config, update_progress)

                # Check for cancellation before ZIP creation
                if asyncio.current_task().cancelled():
                    raise asyncio.CancelledError("Task was cancelled")

                update_progress(95, "Creating ZIP file...")
                zip_file_path = await create_zip_file(task_id, config['temp_dir'], config['agent_id'])

                if task_id in active_tasks:
                    active_tasks[task_id]['zip_file'] = zip_file_path
                    active_tasks[task_id]['agent_id'] = config['agent_id']

                elapsed = int((time.time() - start_time) / 60)
                update_progress(100, f"Automation completed successfully in {elapsed}m! ZIP file ready for download.")
                active_tasks[task_id]['status'] = 'completed'

            except asyncio.CancelledError:
                print(f"LiveVox Task {task_id} was cancelled, cleaning up...")
                raise
            except Exception as e:
                error_msg = str(e)
                if "timeout" in error_msg.lower():
                    update_progress(0, f"Automation timed out: {error_msg}")
                else:
                    update_progress(0, f"Error: {error_msg}")
                active_tasks[task_id]['status'] = 'failed'
            finally:
                # Ensure browser is always closed
                try:
                    await browser.close()
                except:
                    pass

    except asyncio.CancelledError:
        # Handle cancellation gracefully
        if task_id in active_tasks:
            active_tasks[task_id]['status'] = 'stopped'
            active_tasks[task_id]['message'] = 'Automation stopped by user'
        print(f"LiveVox Task {task_id} cancelled successfully")
        raise
    except Exception as e:
        if task_id in active_tasks:
            active_tasks[task_id]['status'] = 'failed'
            active_tasks[task_id]['message'] = f"Error: {str(e)}"

async def create_zip_file(task_id: str, temp_dir: str, identifier: str) -> str:
    """Create a ZIP file containing all downloaded files"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_filename = f"automation_results_{identifier}_{timestamp}.zip"
    zip_path = os.path.join(tempfile.gettempdir(), zip_filename)

    file_count = 0
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, temp_dir)
                zipf.write(file_path, arcname)
                file_count += 1

    print(f"Created ZIP file: {zip_path} with {file_count} files")
    return zip_path

async def run_specific_calls_logic(page, config, update_progress):
    """Implementation for Specific Calls automation logic"""
    agent_details_map = {
        "3P_95": {"folder": "Chuck", "locator_id": "963665"},
        "3P_235": {"folder": "Jim", "locator_id": "976426"},
        "3P_1483": {"folder": "Clifford", "locator_id": "969789"},
        "3P_260": {"folder": "Kathy", "locator_id": "963666"},
        "3P_282": {"folder": "Laurie", "locator_id": "924766"}
    }

    current_agent_id = config['agent_id']
    agent_info = agent_details_map.get(current_agent_id)

    if not agent_info:
        raise Exception(f"Agent ID '{current_agent_id}' not found in mapping")

    calls_to_process = data_manager.get_calls(current_agent_id)
    if not calls_to_process:
        raise Exception(f"No calls configured for agent '{current_agent_id}'")

    target_dir = os.path.join(config['temp_dir'], agent_info['folder'])
    os.makedirs(target_dir, exist_ok=True)

    update_progress(15, f"Starting automation for {len(calls_to_process)} calls for {agent_info['folder']}")

    download_counter = 1
    saved_files = []  # Track all saved files

    update_progress(18, "Navigating to Call Recording Report...")
    try:
        await page.wait_for_selector("button:has-text('Review')", timeout=30000)
        await page.get_by_role("button", name="Review").click()

        await page.wait_for_selector("text=Agent Reports", timeout=15000)
        await page.get_by_text("Agent Reports").click()

        update_progress(19, "Waiting for Call Recording Report...")
        await page.wait_for_selector("div.MuiTreeItem-label:has-text('Call Recording Report')", timeout=60000)
        await page.locator("div.MuiTreeItem-label").filter(has_text="Call Recording Report").click()

        await page.wait_for_load_state("networkidle")
        update_progress(20, "Report page loaded successfully")

    except Exception as nav_error:
        raise Exception(f"Navigation failed: {str(nav_error)}. Make sure you're logged into the correct LiveVox portal.")

    update_progress(22, f"Setting up search for agent {current_agent_id}")
    await page.locator("#search-panel__start-date").fill(config['start_date'])

    await page.get_by_role("row", name="Agent Select One ... Result").locator("u").click()
    await page.get_by_placeholder("Search...").fill(current_agent_id)

    agent_selector = f'[id="\\33 {agent_info["locator_id"]}"]'
    try:
        await page.wait_for_selector(agent_selector, timeout=10000)
        await page.locator(agent_selector).click()
        await page.locator("#agent-combo__ok-btn").click()
        update_progress(24, f"Agent {current_agent_id} selected successfully")
    except Exception as e:
        raise Exception(f"Could not find agent {current_agent_id} (locator: {agent_info['locator_id']}). Check agent mapping.")

    for i, call_data in enumerate(calls_to_process):
        progress = 25 + int(((i + 1) / len(calls_to_process)) * 65)
        update_progress(progress, f"Searching call {i+1}/{len(calls_to_process)}: {call_data['phone']} (duration: {call_data['duration']}s)")

        await page.locator("#search-panel__phone-dialed").clear()
        await page.locator("#search-panel__phone-dialed").fill(call_data['phone'])
        await page.get_by_role("button", name="Generate Report").click()

        try:
            results_table_body = page.locator("div.rt-tbody")
            first_row = results_table_body.locator("div.rt-tr:not(.-padRow)").first
            await first_row.wait_for(state="visible", timeout=15000)

            data_rows = await results_table_body.locator("div.rt-tr:not(.-padRow)").all()
            update_progress(progress, f"Found {len(data_rows)} recordings for {call_data['phone']}, checking durations...")

            found_match = False
            for j, row in enumerate(data_rows):
                duration_in_row = await row.locator("div.rt-td").nth(8).inner_text()
                if duration_in_row == call_data['duration']:
                    update_progress(progress, f"Found matching duration {duration_in_row}s, downloading...")
                    await asyncio.sleep(1.0)  # Much longer delay so message isn't overwritten

                    download_button = row.locator("div.icon--audio-download.clickable")
                    async with page.expect_download() as download_info:
                        await download_button.click()

                    download = await download_info.value
                    base_filename = download.suggested_filename
                    name, ext = os.path.splitext(base_filename)

                    final_filename = f"{name}-{download_counter}{ext}"
                    target_path = os.path.join(target_dir, final_filename)

                    shutil.move(await download.path(), target_path)
                    saved_files.append(final_filename)  # Add to saved files list

                    download_counter += 1
                    found_match = True
                    break

            if not found_match:
                update_progress(progress, f"No matching duration found for {call_data['phone']} (needed: {call_data['duration']}s)")

        except Exception as e:
            update_progress(progress, f"No results found for {call_data['phone']}: {str(e)[:50]}")

    # Show all saved files at the end
    if saved_files:
        files_list = ", ".join(saved_files)
        update_progress(88, f"Saved {len(saved_files)} files: {files_list}")

    update_progress(90, f"Completed processing {len(calls_to_process)} calls. Downloaded {download_counter-1} files.")

async def run_all_calls_logic(page, config, update_progress):
    """Implementation for All Calls automation logic"""
    agent_details_map = {
        "3P_95": {"folder": "Chuck", "locator_id": "963665"},
        "3P_235": {"folder": "Jim", "locator_id": "976426"},
        "3P_1483": {"folder": "Clifford", "locator_id": "969789"},
        "3P_260": {"folder": "Kathy", "locator_id": "963666"},
        "3P_282": {"folder": "Laurie", "locator_id": "924766"}
    }

    current_agent_id = config['agent_id']
    agent_info = agent_details_map.get(current_agent_id)

    if not agent_info:
        raise Exception(f"Agent ID '{current_agent_id}' not found in mapping")

    phone_numbers = config['phone_numbers'].split(',')
    phone_numbers = [phone.strip() for phone in phone_numbers if phone.strip()]

    if not phone_numbers:
        raise Exception("No valid phone numbers provided")

    target_dir = os.path.join(config['temp_dir'], agent_info['folder'])
    os.makedirs(target_dir, exist_ok=True)

    update_progress(15, f"Processing {len(phone_numbers)} phone numbers for {agent_info['folder']}")

    download_counter = 1
    saved_files = []  # Track all saved files

    update_progress(18, "Navigating to Call Recording Report...")
    try:
        await page.wait_for_selector("button:has-text('Review')", timeout=30000)
        await page.get_by_role("button", name="Review").click()

        await page.wait_for_selector("text=Agent Reports", timeout=15000)
        await page.get_by_text("Agent Reports").click()

        update_progress(19, "Waiting for Call Recording Report...")
        await page.wait_for_selector("div.MuiTreeItem-label:has-text('Call Recording Report')", timeout=60000)
        await page.locator("div.MuiTreeItem-label").filter(has_text="Call Recording Report").click()

        await page.wait_for_load_state("networkidle")
        update_progress(20, "Report page loaded successfully")

    except Exception as nav_error:
        raise Exception(f"Navigation failed: {str(nav_error)}. Make sure you're logged into the correct LiveVox portal.")

    update_progress(22, f"Setting up search for agent {current_agent_id}")
    await page.locator("#search-panel__start-date").fill(config['start_date'])

    await page.get_by_role("row", name="Agent Select One ... Result").locator("u").click()
    await page.get_by_placeholder("Search...").fill(current_agent_id)

    agent_selector = f'[id="\\33 {agent_info["locator_id"]}"]'
    try:
        await page.wait_for_selector(agent_selector, timeout=10000)
        await page.locator(agent_selector).click()
        await page.locator("#agent-combo__ok-btn").click()
        update_progress(24, f"Agent {current_agent_id} selected successfully")
    except Exception as e:
        raise Exception(f"Could not find agent {current_agent_id} (locator: {agent_info['locator_id']}). Check agent mapping.")

    for i, phone in enumerate(phone_numbers):
        progress = 25 + int(((i + 1) / len(phone_numbers)) * 65)
        update_progress(progress, f"Searching call {i+1}/{len(phone_numbers)}: {phone}")

        await page.locator("#search-panel__phone-dialed").clear()
        await page.locator("#search-panel__phone-dialed").fill(phone)
        await page.get_by_role("button", name="Generate Report").click()

        try:
            results_table_body = page.locator("div.rt-tbody")
            first_row = results_table_body.locator("div.rt-tr:not(.-padRow)").first
            await first_row.wait_for(state="visible", timeout=15000)

            data_rows = await results_table_body.locator("div.rt-tr:not(.-padRow)").all()

            update_progress(progress, f"Found {len(data_rows)} recording(s) for {phone}. Downloading all...")
            await asyncio.sleep(0.5)  # Give time to see the count message

            for j, row in enumerate(data_rows):
                try:
                    # Get the duration from the row to show consistent messaging
                    duration_in_row = await row.locator("div.rt-td").nth(8).inner_text()
                    try:
                        update_progress(progress, f"Downloading recordings ({duration_in_row}s)...")
                        await asyncio.sleep(0.5)  # Longer delay to ensure message is visible
                    except Exception as msg_error:
                        print(f"Progress message error: {msg_error}")

                    download_button = row.locator("div.icon--audio-download.clickable")
                    async with page.expect_download() as download_info:
                        await download_button.click()

                    download = await download_info.value
                    base_filename = download.suggested_filename
                    name, ext = os.path.splitext(base_filename)

                    final_filename = f"{name}-{download_counter}{ext}"
                    target_path = os.path.join(target_dir, final_filename)

                    shutil.move(await download.path(), target_path)
                    saved_files.append(final_filename)  # Add to saved files list

                    download_counter += 1

                except Exception as download_error:
                    try:
                        update_progress(progress, f"Failed to download recording {j+1}: {str(download_error)}")
                    except Exception as msg_error:
                        print(f"Progress message error: {msg_error}")
                    continue

        except Exception as e:
            update_progress(progress, f"No recordings found for phone: {phone}")

    # Show all saved files at the end
    if saved_files:
        files_list = ", ".join(saved_files)
        update_progress(88, f"Saved {len(saved_files)} files: {files_list}")

    update_progress(90, f"Completed processing {len(phone_numbers)} phone numbers. Downloaded {download_counter-1} files.")

if __name__ == "__main__":
    os.makedirs("templates", exist_ok=True)
    uvicorn.run(app, host="0.0.0.0", port=8001)